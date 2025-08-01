from sqlalchemy import func, text
from app.crud.carelink_crud import CareLinkCrud
from app.database.connection import get_carelink_db
import os
import tempfile
from docxtpl import DocxTemplate
from datetime import datetime
from typing import Optional
from fastapi import Query, status
from app.dto.v1.request.activities import (
    ActividadesGrupalesCreate,
    ActividadesGrupalesUpdate,
)
from app.dto.v1.request.authorized_users import AuthorizedUserUpdate, RoleEnum
from app.dto.v1.request.bill import CalculatePartialBillRequestDTO
from app.dto.v1.request.clinical_evolution import (
    ClinicalEvolutionCreate,
    ClinicalEvolutionUpdate,
)
from app.dto.v1.request.family_member_create_request_dto import (
    AssociateFamilyMemberRequestDTO,
    CreateFamilyMemberRequestDTO,
    UpdateFamilyMemberRequestDTO,
)
from app.dto.v1.response.authorized_user import AuthorizedUser as AuthorizedUserDTO
from app.dto.v1.request.medical_report import ReporteClinicoCreate, ReporteClinicoUpdate
from app.dto.v1.request.payment_method import CreateUserPaymentRequestDTO
from app.dto.v1.request.rates import TarifasServicioUpdateRequestDTO
from app.dto.v1.response.rates import (
    TarifasServicioResponseDTO,
    TarifaServicioResponseDTO,
)
from app.dto.v1.request.user_create_request_dto import AuthorizedUserCreateRequestDTO
from app.dto.v1.request.user_medical_record_create_request_dto import (
    CreateUserAssociatedCaresRequestDTO,
    CreateUserAssociatedInterventionsRequestDTO,
    CreateUserAssociatedMedicinesRequestDTO,
    CreateUserAssociatedVaccinesRequestDTO,
    CreateUserMedicalRecordCreateRequestDTO,
)
from app.dto.v1.request.user_medical_record_update_request_dto import (
    UpdateUserMedicalRecordRequestDTO,
)
from app.dto.v1.request.user_request_dto import UserCreateRequestDTO
from app.dto.v1.response.activities import ActivitiesResponse, TypeOfActivityResponse
from app.dto.v1.response.cares_per_user import (
    CaresPerUserResponseDTO,
    CaresPerUserUpdateDTO,
)
from app.dto.v1.response.clinical_evolution import ClinicalEvolutionResponse
from app.dto.v1.response.create_user import CreateUserResponseDTO
from app.dto.v1.request.user_login_request_dto import UserLoginRequestDTO
from app.dto.v1.response.create_user_medical_record import (
    CreateUserMedicalRecordResponseDTO,
)
from app.dto.v1.response.family_member import FamilyMemberResponseDTO
from app.dto.v1.response.generic_response import Response
from app.dto.v1.response.interventions_per_user import (
    InterventionsPerUserResponseDTO,
    InterventionsPerUserUpdateDTO,
)
from app.dto.v1.response.medical_report import ReporteClinicoResponse
from app.dto.v1.response.medicines_per_user import (
    MedicinesPerUserResponseDTO,
    MedicinesPerUserUpdateDTO,
)
from app.dto.v1.response.payment_method import (
    PaymentMethodResponseDTO,
    PaymentResponseDTO,
    PaymentTypeResponseDTO,
    BillPaymentsTotalResponseDTO,
)
from app.dto.v1.response.professional import ProfessionalResponse
from app.dto.v1.response.user_info import UserInfo
from app.dto.v1.response.user import UserResponseDTO
from app.dto.v1.response.home_visit import (
    VisitaDomiciliariaResponseDTO,
    VisitaDomiciliariaConProfesionalResponseDTO,
)
from app.dto.v1.response.user_flow import UserFlowResponseDTO
from app.dto.v1.response.quarterly_visits import QuarterlyVisitsResponseDTO
from app.dto.v1.response.monthly_payments import MonthlyPaymentsResponseDTO
from app.dto.v1.response.operational_efficiency import OperationalEfficiencyResponseDTO
from app.dto.v1.request.home_visit import (
    VisitaDomiciliariaCreateDTO,
    VisitaDomiciliariaUpdateDTO,
)
from app.dto.v1.response.family_members_by_user import FamilyMembersByUserResponseDTO
from app.dto.v1.response.activity_users import (
    ActivityWithUsersDTO,
    UserForActivityDTO,
    AssignUsersToActivityDTO,
    UpdateUserActivityStatusDTO,
)
from app.dto.v1.response.vaccines_per_user import (
    VaccinesPerUserResponseDTO,
    VaccinesPerUserUpdateDTO,
)
from app.models.activities import ActividadesGrupales
from app.models.authorized_users import AuthorizedUsers
from app.models.cares_per_user import CuidadosEnfermeriaPorUsuario
from app.models.clinical_evolutions import EvolucionesClinicas
from app.models.family_member import FamilyMember
from app.models.family_members_by_user import FamiliaresYAcudientesPorUsuario
from app.models.interventions_per_user import IntervencionesPorUsuario
from app.models.medical_record import MedicalRecord
from app.models.medical_report import ReportesClinicos
from app.models.medicines_per_user import MedicamentosPorUsuario
from app.models.professional import Profesionales
from app.models.user import User
from app.models.contracts import (
    Contratos,
    Facturas,
    MetodoPago,
    Pagos,
    ServiciosPorContrato,
    FechasServicio,
    TipoPago,
    EstadoFactura,
    Servicios,
)
from app.dto.v1.request.contracts import (
    ContratoCreateDTO,
    ContratoUpdateDTO,
    FacturaCreate,
    FacturaCreateWithDetails,
    PagoCreate,
    PagoCreateDTO,
    PagoResponseDTO,
)
from app.dto.v1.request.attendance_schedule import (
    CronogramaAsistenciaCreateDTO,
    CronogramaAsistenciaPacienteCreateDTO,
    CronogramaAsistenciaUpdateDTO,
    EstadoAsistenciaUpdateDTO,
)
from app.dto.v1.response.contracts import (
    ContratoResponseDTO,
    FacturaOut,
    FechaServicioDTO,
    ServicioContratoDTO,
)
from app.dto.v1.response.attendance_schedule import (
    CronogramaAsistenciaResponseDTO,
    CronogramaAsistenciaPacienteResponseDTO,
    PacientePorFechaDTO,
    AsistenciaDiariaResponseDTO,
)

# Nuevos imports para transporte
from app.models.transporte import CronogramaTransporte
from app.dto.v1.request.transport_schedule import (
    CronogramaTransporteCreateDTO,
    CronogramaTransporteUpdateDTO,
)
from app.dto.v1.response.transport_schedule import (
    CronogramaTransporteResponseDTO,
    RutaTransporteResponseDTO,
    RutaDiariaResponseDTO,
)
from app.models.vaccines import VacunasPorUsuario
from app.security.jwt_utilities import (
    decode_access_token,
    create_access_token,
    hash_password,
)
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from functools import lru_cache
from http import HTTPStatus
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, date, time
import json
from app.models.attendance_schedule import (
    CronogramaAsistencia,
    CronogramaAsistenciaPacientes,
)
from app.models.rates import TarifasServicioPorAnio
from app.exceptions.exceptions_classes import EntityNotFoundError
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from enum import Enum


token_auth_scheme = HTTPBearer()
router = APIRouter()


class Role(Enum):
    ADMIN = "admin"
    PROFESSIONAL = "profesional"
    TRANSPORT = "transporte"


@lru_cache()
def get_crud(
    carelink_db: Session = Depends(get_carelink_db),
):
    return CareLinkCrud(carelink_db)


def get_payload(credentials: HTTPAuthorizationCredentials = Depends(token_auth_scheme)):
    payload = decode_access_token(credentials.credentials)
    return payload


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(token_auth_scheme),
    crud: CareLinkCrud = Depends(get_crud),
) -> AuthorizedUsers:
    token = credentials.credentials
    payload = decode_access_token(token)
    user_id = payload.get("sub")
    if user_id is None:
        raise HTTPException(
            status_code=401,
            detail="Credenciales inválidas. Revise sus datos e intente de nuevo.",
        )

    user = crud._get_authorized_user_info(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


def require_role(required_role: str):
    def role_checker(user: dict = Depends(get_current_user)):
        if user.role != required_role:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tiene el rol requerido para realizar esta acción",
            )
        return user

    return role_checker


def require_roles(*allowed_roles: str):
    def role_checker(user: dict = Depends(get_current_user)):
        if user.role not in allowed_roles:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No tienes los suficientes permisos",
            )
        return user

    return role_checker


@router.get("/users", response_model=Response[List[UserResponseDTO]])
async def list_users(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[UserResponseDTO]]:
    users = crud.list_users_without_home_visits()
    users_list = []
    for user in users:
        users_list.append(user.__dict__)
    return Response[List[UserResponseDTO]](
        data=users_list,
        status_code=HTTPStatus.OK,
        message="Usuarios consultados con éxito",
        error=None,
    )


@router.get("/users/home-visits", response_model=Response[List[UserResponseDTO]])
async def list_users_with_home_visits(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[UserResponseDTO]]:
    users = crud.list_users_with_home_visits()
    users_list = []
    for user in users:
        users_list.append(user.__dict__)
    return Response[List[UserResponseDTO]](
        data=users_list,
        status_code=HTTPStatus.OK,
        message="Usuarios con visitas domiciliarias consultados con éxito",
        error=None,
    )


@router.get("/users/{id}", response_model=Response[UserResponseDTO])
async def list_user_by_id(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[UserResponseDTO]:
    user = crud.list_user_by_user_id(id)
    user_response = UserResponseDTO(**user.__dict__)
    return Response[UserResponseDTO](
        data=user_response,
        status_code=HTTPStatus.OK,
        message="Usuario consultado con éxito",
        error=None,
    )


@router.get(
    "/family_members",
    status_code=200,
    response_model=Response[List[FamilyMemberResponseDTO]],
)
async def get_family_members(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    family_members = crud._get_family_members()
    family_members_dto = [
        FamilyMemberResponseDTO.from_orm(member) for member in family_members
    ]
    return Response[List[FamilyMemberResponseDTO]](
        data=family_members_dto,
        message="Acudientes consultados exitosamente",
        error=None,
        status_code=HTTPStatus.OK,
    )


@router.get(
    "/family_members/{id}",
    status_code=200,
    response_model=Response[FamilyMemberResponseDTO],
)
async def get_family_member_by_id(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    family_member, parentesco = crud._get_family_member_by_id(id)
    family_member.parentesco = parentesco
    family_member_dto = FamilyMemberResponseDTO.from_orm(family_member)

    return Response[FamilyMemberResponseDTO](
        data=family_member_dto,
        message="Acudiente consultado exitosamente",
        error=None,
        status_code=HTTPStatus.OK,
    )


@router.get(
    "/users/{id}/family_members",
    status_code=200,
    response_model=Response[List[FamilyMembersByUserResponseDTO]],
)
async def get_user_family_members(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    family_members = crud._get_family_members_by_user_id(id)
    family_members_dto = [
        FamilyMembersByUserResponseDTO.from_orm(member) for member in family_members
    ]
    return Response[List[FamilyMembersByUserResponseDTO]](
        data=family_members_dto,
        message="Acudientes consultados de manera exitosa",
        error=None,
        status_code=HTTPStatus.OK,
    )


@router.get(
    "/users/{id}/medical_record",
    status_code=200,
    response_model=Response[CreateUserMedicalRecordResponseDTO | None],
)
async def get_user_medical_record(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    medical_record = crud.list_user_medical_record(id)
    medical_record_response = None
    if medical_record:
        medical_record_response = CreateUserMedicalRecordResponseDTO(
            **medical_record.__dict__
        )

    return Response[CreateUserMedicalRecordResponseDTO | None](
        data=medical_record_response,
        message="Historia clínica listada",
        error=None,
        status_code=200,
    )


@router.get(
    "/record/{id}/medicines",
    status_code=200,
    response_model=Response[List[MedicinesPerUserResponseDTO]],
)
async def get_record_medicines(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    medicines = crud._get_user_medicines_by_medical_record_id(id)
    result = [
        MedicinesPerUserResponseDTO(**medicine.__dict__) for medicine in medicines
    ]
    return Response[List[MedicinesPerUserResponseDTO]](
        data=result,
        message="Medicamentos consultados con éxito",
        status_code=201,
        error=None,
    )


@router.get(
    "/record/{id}/interventions",
    status_code=200,
    response_model=Response[List[InterventionsPerUserResponseDTO]],
)
async def get_record_interventions(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    interventions = crud._get_user_interventions_by_medical_record_id(id)
    result = [
        InterventionsPerUserResponseDTO(**intervention.__dict__)
        for intervention in interventions
    ]
    return Response[List[InterventionsPerUserResponseDTO]](
        data=result, message="Success", status_code=201, error=None
    )


@router.get(
    "/record/{id}/cares",
    status_code=200,
    response_model=Response[List[CaresPerUserResponseDTO]],
)
async def get_record_cares(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    cares = crud._get_user_cares_by_medical_record_id(id)
    result = [CaresPerUserResponseDTO(**care.__dict__) for care in cares]
    return Response[List[CaresPerUserResponseDTO]](
        data=result, message="Success", status_code=201, error=None
    )


@router.get(
    "/record/{id}/vaccines",
    status_code=200,
    response_model=Response[List[VaccinesPerUserResponseDTO]],
)
async def get_record_vaccines(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    vaccines = crud._get_user_vaccines_by_medical_record_id(id)
    result = [VaccinesPerUserResponseDTO(**vaccine.__dict__) for vaccine in vaccines]
    return Response[List[VaccinesPerUserResponseDTO]](
        data=result, message="Success", status_code=201, error=None
    )


@router.get("/info", status_code=200, response_model=Response[UserInfo])
async def get_user_info(
    crud: CareLinkCrud = Depends(get_crud), payload: dict = Depends(get_payload)
):
    user = crud._get_authorized_user_info(payload["sub"])
    result = UserInfo(**user.__dict__)
    return Response[UserInfo](
        data=result,
        message="User info retrieved successfuly",
        error=None,
        status_code=HTTPStatus.OK,
    )


@router.get(
    "/medical_reports/{reporte_id}", response_model=Response[ReporteClinicoResponse]
)
def get_reporte_clinico(
    reporte_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ReporteClinicoResponse]:
    report = crud._get_medical_report_by_id(reporte_id)
    report_dict = report.__dict__
    professional_response = None
    if report.profesional:
        professional_response = ProfessionalResponse.from_orm(report.profesional)
    report_response = ReporteClinicoResponse(**report_dict)
    del report_response.profesional
    report_response.profesional = professional_response
    return Response[ReporteClinicoResponse](
        data=report_response,
        message="Reporte consultado",
        status_code=200,
        error=None,
    )


@router.get(
    "/user/{user_id}/medical_reports",
    response_model=Response[List[ReporteClinicoResponse]],
)
def get_medical_reports(
    user_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[ReporteClinicoResponse]]:
    reports = crud._get_medical_reports_by_user_id(user_id)
    return Response[List[ReporteClinicoResponse]](
        data=[ReporteClinicoResponse.from_orm(report) for report in reports],
        message="Reporte consultado",
        status_code=200,
        error=None,
    )


@router.get("/professionals", response_model=Response[List[ProfessionalResponse]])
async def get_professionals(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[ProfessionalResponse]]:
    professionals = crud._get_professionals()
    response = [
        ProfessionalResponse.from_orm(professional) for professional in professionals
    ]
    return Response[List[ProfessionalResponse]](
        data=response, message="Lista de profesionales", status_code=200, error=None
    )


@router.get(
    "/reports/{id}/evolutions", response_model=Response[List[ClinicalEvolutionResponse]]
)
async def get_clinical_evolutions(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[ClinicalEvolutionResponse]]:
    evolutions = crud._get_clinical_evolutions_by_report_id(id)
    response = [
        ClinicalEvolutionResponse.from_orm(evolution) for evolution in evolutions
    ]
    return Response[List[ClinicalEvolutionResponse]](
        data=response,
        message="Evoluciones clínicas consultadas con éxito",
        error=None,
        status_code=200,
    )


@router.get("/evolutions/{id}", response_model=Response[ClinicalEvolutionResponse])
async def get_clinical_evolution(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ClinicalEvolutionResponse]:
    evolution = crud._get_clinical_evolution_by_evolution_id(id)
    response = ClinicalEvolutionResponse.from_orm(evolution)

    return Response[ClinicalEvolutionResponse](
        data=response,
        message="Evoluciones clínicas consultadas con éxito",
        error=None,
        status_code=200,
    )


@router.get(
    "/activities",
    status_code=200,
    response_model=Response[List[ActivitiesResponse]],
)
async def get_activities(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[ActivitiesResponse]]:
    activities = crud._get_activities()
    result = [ActivitiesResponse(**activity.__dict__) for activity in activities]
    return Response[List[ActivitiesResponse]](
        data=result,
        message="Actividades consultadas con éxito",
        status_code=201,
        error=None,
    )


@router.get(
    "/activities/{id}",
    status_code=200,
    response_model=Response[ActivitiesResponse],
)
async def get_activity_by_id(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ActivitiesResponse]:
    activity = crud._get_activity_by_id(id)
    result = ActivitiesResponse(**activity.__dict__)
    return Response[ActivitiesResponse](
        data=result,
        message="Actividad consultada con éxito",
        status_code=201,
        error=None,
    )


@router.get(
    "/activity_types",
    status_code=200,
    response_model=Response[List[TypeOfActivityResponse]],
)
async def get_activity_types(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[TypeOfActivityResponse]]:
    activity_types = crud._get_activity_types()
    result = [TypeOfActivityResponse(**type.__dict__) for type in activity_types]
    return Response[List[TypeOfActivityResponse]](
        data=result,
        message="Tipos de actividad consultados con éxito",
        status_code=201,
        error=None,
    )


@router.get("/metodos_pago", response_model=Response[List[PaymentMethodResponseDTO]])
async def get_payment_methods(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[PaymentMethodResponseDTO]]:
    payment_methods = crud._get_payment_methods()
    payment_methods_response = [
        PaymentMethodResponseDTO.from_orm(method) for method in payment_methods
    ]
    return Response[List[PaymentMethodResponseDTO]](
        data=payment_methods_response,
        status_code=HTTPStatus.OK,
        error=None,
        message="Métodos de pago retornados con éxito",
    )


@router.get("/tipos_pago", response_model=Response[List[PaymentTypeResponseDTO]])
async def get_payment_types(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[PaymentTypeResponseDTO]]:
    payment_types = crud._get_payment_types()
    payment_types_response = [
        PaymentTypeResponseDTO.from_orm(type_pago) for type_pago in payment_types
    ]
    return Response[List[PaymentTypeResponseDTO]](
        data=payment_types_response,
        status_code=HTTPStatus.OK,
        error=None,
        message="Tipos de pago retornados con éxito",
    )


@router.get(
    "/activities-upcoming",
    status_code=200,
    response_model=Response[List[ActivitiesResponse]],
)
async def get_upcoming_activities(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[ActivitiesResponse]]:
    activities = crud._get_upcoming_activities()
    result = [ActivitiesResponse(**activity.__dict__) for activity in activities]
    return Response[List[ActivitiesResponse]](
        data=result,
        message="Actividades consultadas con éxito",
        status_code=201,
        error=None,
    )


@router.post("/pagos/registrar", response_model=Response[PaymentResponseDTO])
async def register_payment(
    payment: CreateUserPaymentRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[PaymentResponseDTO]:
    try:
        # Validar que la factura existe
        try:
            bill = crud.get_bill_by_id(payment.id_factura)
        except EntityNotFoundError:
            raise HTTPException(
                status_code=400,
                detail=f"La factura con ID {payment.id_factura} no existe",
            )

        # Validar que el método de pago existe
        payment_methods = crud._get_payment_methods()
        if not any(
            pm.id_metodo_pago == payment.id_metodo_pago for pm in payment_methods
        ):
            raise HTTPException(
                status_code=400,
                detail=f"El método de pago con ID {payment.id_metodo_pago} no existe",
            )

        # Validar que el tipo de pago existe
        payment_types = crud._get_payment_types()
        if not any(pt.id_tipo_pago == payment.id_tipo_pago for pt in payment_types):
            raise HTTPException(
                status_code=400,
                detail=f"El tipo de pago con ID {payment.id_tipo_pago} no existe",
            )

        # Crear el objeto de pago
        payment_data = Pagos(
            id_factura=payment.id_factura,
            id_metodo_pago=payment.id_metodo_pago,
            id_tipo_pago=payment.id_tipo_pago,
            fecha_pago=payment.fecha_pago,
            valor=payment.valor,
        )

        # Crear el pago usando el CRUD
        payment_response = crud.create_payment(payment_data)

        # Actualizar el estado de la factura según los pagos
        crud.update_factura_status(payment.id_factura)

        # Actualizar el estado de la factura según los pagos
        crud.update_factura_status(payment.id_factura)

        # Actualizar el estado de la factura según los pagos
        crud.update_factura_status(payment.id_factura)

        # Actualizar el estado de la factura según los pagos
        crud.update_factura_status(payment.id_factura)

        return Response[PaymentResponseDTO](
            data=PaymentResponseDTO.from_orm(payment_response),
            error=None,
            message="Pago creado de manera exitosa",
            status_code=HTTPStatus.CREATED,
        )

    except HTTPException:
        # Re-lanzar HTTPExceptions para mantener el status code correcto
        raise
    except Exception as e:
        # Log del error para debugging
        print(f"Error inesperado en register_payment: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.post("/calcular/factura", response_model=Response[float])
async def calculate_partial_bill(
    partial_bill: CalculatePartialBillRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[float]:
    try:
        total = crud.calculate_partial_bill(
            partial_bill.service_ids, partial_bill.quantities, partial_bill.year
        )
        return Response[float](
            data=total,
            message="Factura calculada con éxito",
            status_code=HTTPStatus.OK,
            error=None,
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al calcular factura: {str(e)}"
        )


@router.post("/calcular/total_factura", response_model=Response[float])
async def calculate_total_factura(
    payload: dict,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[float]:
    """
    Calcula el total de factura incluyendo impuestos y descuentos

    Args:
        payload: Diccionario con subtotal, impuestos y descuentos
        crud: Instancia del CRUD

    Returns:
        Response con el total calculado

    Raises:
        HTTPException: Si hay errores en el cálculo
    """
    try:
        subtotal = float(payload.get("subtotal", 0))
        impuestos = float(payload.get("impuestos", 0))
        descuentos = float(payload.get("descuentos", 0))

        # Validar que los valores no sean negativos
        if subtotal < 0:
            raise HTTPException(
                status_code=400, detail="El subtotal no puede ser negativo"
            )
        if impuestos < 0:
            raise HTTPException(
                status_code=400, detail="Los impuestos no pueden ser negativos"
            )
        if descuentos < 0:
            raise HTTPException(
                status_code=400, detail="Los descuentos no pueden ser negativos"
            )

        # Calcular total: subtotal + impuestos - descuentos
        total_factura = subtotal + impuestos - descuentos

        # Asegurar que el total no sea negativo
        if total_factura < 0:
            total_factura = 0

        return Response[float](
            data=total_factura,
            message="Total de factura calculado correctamente",
            status_code=HTTPStatus.OK,
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al calcular total de factura: {str(e)}"
        )


@router.post("/users", status_code=201, response_model=Response[dict])
async def create_users(
    user: str = Form(...),
    photo: Optional[UploadFile] = File(None),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[dict]:
    """
    Crea un nuevo usuario en el sistema.

    Si el campo 'visitas_domiciliarias' es True, también crea automáticamente
    un registro en la tabla VisitasDomiciliarias con los datos del usuario.

    Args:
        user: Datos del usuario en formato JSON
        photo: Archivo de imagen opcional para la foto del usuario

    Returns:
        Response con los datos del usuario creado y opcionalmente
        los datos de la visita domiciliaria si fue creada.
    """
    try:
        user_data = UserCreateRequestDTO.parse_raw(user)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid user data: {str(e)}")

    user_to_save = User(**user_data.dict())

    saved_user = crud.save_user(user_to_save, photo)

    # Si el usuario requiere visitas domiciliarias, crear el registro correspondiente
    home_visit_response = None
    if saved_user.visitas_domiciliarias:
        user_dict = user_data.dict()
        home_visit = crud.create_home_visit(saved_user.id_usuario, user_dict)
        if home_visit:  # Solo crear respuesta si se creó la visita
            home_visit_response = VisitaDomiciliariaResponseDTO.from_orm(home_visit)

    # Crear manualmente el diccionario con los campos necesarios
    user_dict = {
        "id_usuario": saved_user.id_usuario,
        "apellidos": saved_user.apellidos,
        "direccion": saved_user.direccion,
        "email": saved_user.email,
        "escribe": saved_user.escribe,
        "estado": saved_user.estado,
        "estado_civil": saved_user.estado_civil,
        "fecha_nacimiento": saved_user.fecha_nacimiento,
        "fecha_registro": saved_user.fecha_registro,
        "genero": saved_user.genero,
        "grado_escolaridad": saved_user.grado_escolaridad,
        "ha_estado_en_otro_centro": saved_user.ha_estado_en_otro_centro,
        "lee": saved_user.lee,
        "lugar_nacimiento": saved_user.lugar_nacimiento,
        "lugar_procedencia": saved_user.lugar_procedencia,
        "n_documento": saved_user.n_documento,
        "nombres": saved_user.nombres,
        "nucleo_familiar": saved_user.nucleo_familiar,
        "ocupacion_quedesempeño": saved_user.ocupacion_quedesempeño,
        "origen_otrocentro": saved_user.origen_otrocentro,
        "proteccion_exequial": saved_user.proteccion_exequial,
        "regimen_seguridad_social": saved_user.regimen_seguridad_social,
        "telefono": saved_user.telefono,
        "tipo_afiliacion": saved_user.tipo_afiliacion,
        "url_imagen": saved_user.url_imagen,
        "profesion": saved_user.profesion,
        "tipo_usuario": saved_user.tipo_usuario,
        "visitas_domiciliarias": saved_user.visitas_domiciliarias,
    }
    user_response = UserResponseDTO(**user_dict)

    # Preparar la respuesta con información adicional si se creó visita domiciliaria
    response_data = {"user": user_response, "home_visit": home_visit_response}

    message = "Usuario creado de manera exitosa"
    if home_visit_response:
        message += " con visita domiciliaria programada"

    return Response[dict](
        data=response_data,
        status_code=HTTPStatus.CREATED,
        message=message,
        error=None,
    )


@router.post(
    "/family_members/:id",
    status_code=201,
    response_model=Response[object],
)
async def create_family_members(
    id: int,
    family_member: CreateFamilyMemberRequestDTO,
    kinship: AssociateFamilyMemberRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    try:
        family_member_to_save = FamilyMember(**family_member.dict())
        crud.save_family_member(id, kinship, family_member_to_save)

        return Response[object](
            data={},
            status_code=HTTPStatus.CREATED,
            message="Acudiente registrado de manera exitosa",
            error=None,
        )
    except ValueError as e:
        raise HTTPException(status_code=HTTPStatus.BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            detail=f"Error interno del servidor: {str(e)}",
        )


@router.post("/login", response_model=Response[dict])
async def login_user(
    login_data: UserLoginRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
) -> Response[dict]:
    user = crud.authenticate_user(login_data.email, login_data.password)
    if not user:
        raise HTTPException(
            status_code=HTTPStatus.UNAUTHORIZED,
            detail="Usuario o contraseña incorrectos. Revise sus datos e intente de nuevo.",
        )

    access_token = create_access_token(
        data={"sub": str(user.id), "role": str(user.role)}
    )
    return Response[dict](
        data={
            "access_token": access_token,
            "token_type": "bearer",
            "role": str(user.role),
        },
        status_code=HTTPStatus.OK,
        message="Inicio de sesión exitoso",
        error=None,
    )


@router.post("/users/{id}/record", status_code=201, response_model=Response[object])
async def create_user_record(
    id: int,
    record: str = Form(...),
    medicines: str = Form(None),
    cares: str = Form(None),
    interventions: str = Form(None),
    vaccines: str = Form(None),
    attachments: Optional[List[UploadFile]] = File(None),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    try:
        record_data = json.loads(record)
        medicines_data = json.loads(medicines) if medicines else []
        cares_data = json.loads(cares) if cares else []
        interventions_data = json.loads(interventions) if interventions else []
        vaccines_data = json.loads(vaccines) if vaccines else []

        record_to_save = MedicalRecord(**record_data)
        medicines_to_save = [
            MedicamentosPorUsuario(**medicine) for medicine in medicines_data
        ]
        cares_to_save = [CuidadosEnfermeriaPorUsuario(**care) for care in cares_data]
        interventions_to_save = [
            IntervencionesPorUsuario(**intervention)
            for intervention in interventions_data
        ]
        vaccines_to_save = [VacunasPorUsuario(**vaccine) for vaccine in vaccines_data]

        crud.save_user_medical_record(
            id,
            record_to_save,
            medicines_to_save,
            cares_to_save,
            interventions_to_save,
            vaccines_to_save,
            attachments,
        )

        return Response[object](
            data={},
            status_code=HTTPStatus.CREATED,
            message="Historia clínica registrada de manera exitosa",
            error=None,
        )
    except Exception as e:
        print(f"Error creating medical record: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.post(
    "/users/{id}/medical_record",
    status_code=201,
    response_model=Response[CreateUserMedicalRecordResponseDTO],
)
async def create_user_medical_record(
    id: int,
    record: CreateUserMedicalRecordCreateRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[CreateUserMedicalRecordResponseDTO]:
    record_to_save = MedicalRecord(**record.dict())
    saved_record = crud.create_user_medical_record(id, record_to_save)
    response_data = CreateUserMedicalRecordResponseDTO(**saved_record.__dict__)
    return Response[CreateUserMedicalRecordResponseDTO](
        data=response_data,
        message="Historia clínica creada con éxito",
        status_code=201,
        error=None,
    )


@router.get(
    "/professional/{user_id}",
    status_code=201,
    response_model=Response[ProfessionalResponse],
)
async def list_professional_by_user_id(
    user_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ProfessionalResponse]:
    professional = crud._get_professional_by_user_id(user_id)
    professional_response = ProfessionalResponse.from_orm(professional)
    return Response[ProfessionalResponse](
        message="Profesional consultado de manera exitosa",
        status_code=HTTPStatus.OK,
        error=None,
        data=professional_response,
    )


@router.post("/create", status_code=201, response_model=Response[AuthorizedUserDTO])
async def create_user(
    user: AuthorizedUserCreateRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(require_roles(Role.ADMIN.value)),
) -> Response[AuthorizedUserDTO]:
    hashed_password = hash_password(user.password)
    user_to_save = AuthorizedUsers(**user.dict(exclude={"professional_user"}))
    user_to_save.password = hashed_password
    user_to_save.is_deleted = False
    saved_user = crud.create_user(user_to_save)

    if user.role == RoleEnum.profesional:
        if not user.professional_user:
            raise HTTPException(
                status_code=422,
                detail="Professional data is required for professional role.",
            )
        professional_data = user.professional_user
        professional_to_save = Profesionales(
            id_user=saved_user.id,
            direccion=professional_data.home_address,
            apellidos=professional_data.last_name,
            cargo=professional_data.charge,
            e_mail=professional_data.email,
            especialidad=professional_data.specialty,
            estado="Activo",
            fecha_ingreso=professional_data.entry_date,
            fecha_nacimiento=professional_data.birthdate,
            n_documento=professional_data.document_number,
            nombres=professional_data.first_name,
            profesion=professional_data.profession,
            t_profesional=professional_data.professional_id_number,
            telefono=professional_data.phone_number,
        )
        crud.create_professional_user(professional_to_save)

    access_token = create_access_token(
        data={"sub": str(saved_user.id), "role": str(saved_user.role)}
    )
    saved_user.token = access_token

    return Response[AuthorizedUserDTO](
        data=saved_user.__dict__,
        status_code=HTTPStatus.CREATED,
        message="User created successfully",
        error=None,
    )


@router.post("/medical_reports/", response_model=Response[ReporteClinicoResponse])
def create_reporte_clinico(
    id_historiaclinica: int = Form(...),
    id_profesional: int = Form(...),
    IMC: Optional[float] = Form(None),
    Obs_habitosalimenticios: Optional[str] = Form(None),
    Porc_grasacorporal: Optional[float] = Form(None),
    Porc_masamuscular: Optional[float] = Form(None),
    area_afectiva: Optional[str] = Form(None),
    area_comportamental: Optional[str] = Form(None),
    areacognitiva: Optional[str] = Form(None),
    areainterpersonal: Optional[str] = Form(None),
    areasomatica: Optional[str] = Form(None),
    circunferencia_cadera: Optional[float] = Form(None),
    circunferencia_cintura: Optional[float] = Form(None),
    consumo_aguadiaria: Optional[float] = Form(None),
    diagnostico: Optional[str] = Form(None),
    fecha_registro: Optional[str] = Form(None),
    frecuencia_actividadfisica: Optional[str] = Form(None),
    frecuencia_cardiaca: Optional[int] = Form(None),
    frecuencia_respiratoria: Optional[int] = Form(None),
    motivo_consulta: Optional[str] = Form(None),
    nivel_dolor: Optional[int] = Form(None),
    observaciones: Optional[str] = Form(None),
    peso: Optional[int] = Form(None),
    presion_arterial: Optional[int] = Form(None),
    pruebas_examenes: Optional[str] = Form(None),
    recomendaciones: Optional[str] = Form(None),
    remision: Optional[str] = Form(None),
    saturacionOxigeno: Optional[int] = Form(None),
    temperatura_corporal: Optional[float] = Form(None),
    tipo_reporte: Optional[str] = Form(None),
    attachments: Optional[List[UploadFile]] = File(None),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ReporteClinicoResponse]:
    report_data = {
        "id_historiaclinica": id_historiaclinica,
        "id_profesional": id_profesional,
        "IMC": IMC,
        "Obs_habitosalimenticios": Obs_habitosalimenticios,
        "Porc_grasacorporal": Porc_grasacorporal,
        "Porc_masamuscular": Porc_masamuscular,
        "area_afectiva": area_afectiva,
        "area_comportamental": area_comportamental,
        "areacognitiva": areacognitiva,
        "areainterpersonal": areainterpersonal,
        "areasomatica": areasomatica,
        "circunferencia_cadera": circunferencia_cadera,
        "circunferencia_cintura": circunferencia_cintura,
        "consumo_aguadiaria": consumo_aguadiaria,
        "diagnostico": diagnostico,
        "fecha_registro": fecha_registro,
        "frecuencia_actividadfisica": frecuencia_actividadfisica,
        "frecuencia_cardiaca": frecuencia_cardiaca,
        "frecuencia_respiratoria": frecuencia_respiratoria,
        "motivo_consulta": motivo_consulta,
        "nivel_dolor": nivel_dolor,
        "observaciones": observaciones,
        "peso": peso,
        "presion_arterial": presion_arterial,
        "pruebas_examenes": pruebas_examenes,
        "recomendaciones": recomendaciones,
        "remision": remision,
        "saturacionOxigeno": saturacionOxigeno,
        "temperatura_corporal": temperatura_corporal,
        "tipo_reporte": tipo_reporte,
    }

    report_to_save = ReportesClinicos(**report_data)
    resulting_report = crud.save_medical_report(report_to_save, attachments)
    return Response[ReporteClinicoResponse](
        data=ReporteClinicoResponse(**resulting_report.__dict__),
        message="Reporte clínico creado de manera exitosa",
        status_code=200,
        error=None,
    )


@router.post(
    "/reports/{id}/evolutions", response_model=Response[ClinicalEvolutionResponse]
)
def create_clinical_evolution(
    evolution: ClinicalEvolutionCreate,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ClinicalEvolutionResponse]:
    report_to_save = EvolucionesClinicas(**evolution.dict())
    resulting_report = crud.save_clinical_evolution(report_to_save)
    return Response[ClinicalEvolutionResponse](
        data=ClinicalEvolutionResponse(**resulting_report.__dict__),
        message="Evolución clínica creada de manera exitosa",
        status_code=200,
        error=None,
    )


@router.post("/activities", response_model=Response[ActivitiesResponse])
def create_activity(
    activity: ActividadesGrupalesCreate,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ActivitiesResponse]:
    activity_to_save = ActividadesGrupales(**activity.dict())
    resulting_activity = crud.save_activity(activity_to_save)
    return Response[ActivitiesResponse](
        data=ActivitiesResponse(**resulting_activity.__dict__),
        message="Actividad creada de manera exitosa",
        status_code=200,
        error=None,
    )


@router.patch("/users/{id}", status_code=200, response_model=Response[UserResponseDTO])
async def update_user(
    id: int,
    user: str = Form(...),
    photo: Optional[UploadFile] = File(None),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    try:
        user_data = UserCreateRequestDTO.parse_raw(user)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid user data: {str(e)}")

    user_to_save = User(**user_data.dict())

    saved_user = crud.update_user(id, user_to_save, photo)
    user_response = UserResponseDTO(**saved_user.__dict__)

    return Response[UserResponseDTO](
        data=user_response,
        status_code=HTTPStatus.OK,
        message="Usuario actualizado de manera exitosa",
        error=None,
    )


@router.patch("/user/treatment/{id}", status_code=200, response_model=Response[object])
async def update_treatment(
    id: int,
    treatment: MedicinesPerUserUpdateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    treatment_to_update = MedicamentosPorUsuario(**treatment.__dict__)
    crud.update_medical_treatment(id, treatment_to_update)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Medicamento actualizado de manera exitosa",
        error=None,
    )


@router.patch("/user/nursing/{id}", status_code=200, response_model=Response[object])
async def update_nursing(
    id: int,
    treatment: CaresPerUserUpdateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    treatment_to_update = CuidadosEnfermeriaPorUsuario(**treatment.__dict__)
    crud.update_medical_nursing(id, treatment_to_update)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Tratamiento actualizado de manera exitosa",
        error=None,
    )


@router.patch(
    "/user/intervention/{id}", status_code=200, response_model=Response[object]
)
async def update_intervention(
    id: int,
    treatment: InterventionsPerUserUpdateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    treatment_to_update = IntervencionesPorUsuario(**treatment.__dict__)
    crud.update_medical_intervention(id, treatment_to_update)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Intervención actualizada de manera exitosa",
        error=None,
    )


@router.patch("/user/vaccine/{id}", status_code=200, response_model=Response[object])
async def update_vaccine(
    id: int,
    treatment: VaccinesPerUserUpdateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    treatment_to_update = VacunasPorUsuario(**treatment.__dict__)
    crud.update_medical_vaccine(id, treatment_to_update)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Vacuna actualizada de manera exitosa",
        error=None,
    )


@router.patch(
    "/family_members/{id}",
    status_code=200,
    response_model=Response[FamilyMemberResponseDTO],
)
async def update_family_member(
    id: int,
    family_member_id: int,
    family_member: UpdateFamilyMemberRequestDTO,
    kinship: AssociateFamilyMemberRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    try:
        db_family_member = crud._get_family_member_by_id(family_member_id)
        if not db_family_member:
            raise HTTPException(
                status_code=HTTPStatus.NOT_FOUND, detail="Acudiente no encontrado"
            )

        family_member_to_update = FamilyMember(**family_member.dict())

        family_member_updated = crud._update_family_member(
            user_id=id,
            family_member=family_member_to_update,
            kinship=kinship,
            db_family_member=db_family_member[0],
        )

        return Response[FamilyMemberResponseDTO](
            data=family_member_updated.__dict__,
            status_code=HTTPStatus.OK,
            message="Acudiente actualizado de manera exitosa",
            error=None,
        )
    except ValueError as e:
        raise HTTPException(status_code=HTTPStatus.BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            detail=f"Error interno del servidor: {str(e)}",
        )


@router.patch(
    "/users/{id}/medical_record/{record_id}",
    status_code=200,
    response_model=Response[object],
)
async def update_user_medical_record(
    id: int,
    record_id: int,
    record: str = Form(...),
    medicines: str = Form(...),
    cares: str = Form(...),
    interventions: str = Form(...),
    vaccines: str = Form(...),
    attachments: Optional[List[UploadFile]] = File(None),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    import json
    from pydantic import ValidationError
    from io import BytesIO

    try:
        # Parsear los JSON strings con mejor manejo de errores
        try:
            record_data = json.loads(record)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400, detail=f"Error al parsear JSON del record: {str(e)}"
            )

        try:
            medicines_data = json.loads(medicines)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400, detail=f"Error al parsear JSON de medicines: {str(e)}"
            )

        try:
            cares_data = json.loads(cares)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400, detail=f"Error al parsear JSON de cares: {str(e)}"
            )

        try:
            interventions_data = json.loads(interventions)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error al parsear JSON de interventions: {str(e)}",
            )

        try:
            vaccines_data = json.loads(vaccines)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400, detail=f"Error al parsear JSON de vaccines: {str(e)}"
            )

        # Convertir a objetos Pydantic con mejor manejo de errores
        try:
            record_obj = UpdateUserMedicalRecordRequestDTO(**record_data)
        except ValidationError as e:
            raise HTTPException(
                status_code=400, detail=f"Error de validación en record: {str(e)}"
            )

        try:
            medicines_objs = [
                CreateUserAssociatedMedicinesRequestDTO(**medicine)
                for medicine in medicines_data
            ]
        except ValidationError as e:
            raise HTTPException(
                status_code=400, detail=f"Error de validación en medicines: {str(e)}"
            )

        try:
            cares_objs = [
                CreateUserAssociatedCaresRequestDTO(**care) for care in cares_data
            ]
        except ValidationError as e:
            raise HTTPException(
                status_code=400, detail=f"Error de validación en cares: {str(e)}"
            )

        try:
            interventions_objs = [
                CreateUserAssociatedInterventionsRequestDTO(**intervention)
                for intervention in interventions_data
            ]
        except ValidationError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error de validación en interventions: {str(e)}",
            )

        try:
            vaccines_objs = [
                CreateUserAssociatedVaccinesRequestDTO(**vaccine)
                for vaccine in vaccines_data
            ]
        except ValidationError as e:
            raise HTTPException(
                status_code=400, detail=f"Error de validación en vaccines: {str(e)}"
            )

        # Manejar archivos adjuntos
        attachment_urls = []
        if attachments:
            for attachment in attachments:
                if attachment and attachment.filename:
                    # Generar nombre único para el archivo
                    import uuid
                    import os
                    from datetime import datetime

                    file_extension = os.path.splitext(attachment.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"

                    # Crear ruta en S3
                    s3_path = f"medical_records/{id}/{record_id}/{unique_filename}"

                    # Subir archivo a S3
                    try:
                        file_content = await attachment.read()
                        file_obj = BytesIO(file_content)
                        s3_url = crud.upload_file_to_s3(
                            file_obj, "images-care-link", s3_path
                        )
                        attachment_urls.append(s3_url)
                    except Exception as e:
                        print(f"Error al subir archivo a S3: {e}")
                        continue

        # Actualizar el record con las URLs de los archivos adjuntos
        update_data = record_obj.dict(exclude_unset=True)
        if attachment_urls:
            update_data["url_hc_adjunto"] = ",".join(attachment_urls)

        medicines_to_save = [
            MedicamentosPorUsuario(**medicine.__dict__) for medicine in medicines_objs
        ]
        cares_to_save = [
            CuidadosEnfermeriaPorUsuario(**care.__dict__) for care in cares_objs
        ]
        interventions_to_save = [
            IntervencionesPorUsuario(**intervention.__dict__)
            for intervention in interventions_objs
        ]
        vaccines_to_save = [
            VacunasPorUsuario(**vaccine.__dict__) for vaccine in vaccines_objs
        ]

        crud.update_user_medical_record(
            id,
            record_id,
            update_data,
            medicines_to_save,
            cares_to_save,
            interventions_to_save,
            vaccines_to_save,
        )

        return Response[object](
            data={},
            message="Historia clínica actualizada con éxito",
            status_code=200,
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error inesperado en update_user_medical_record: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.patch(
    "/users/{id}/medical_record/{record_id}/simplified",
    status_code=200,
    response_model=Response[object],
)
async def update_user_medical_record_simplified(
    id: int,
    record_id: int,
    record: UpdateUserMedicalRecordRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[object]:
    """Endpoint para actualizar historias clínicas simplificadas (solo el registro principal)"""
    update_data = record.dict(exclude_unset=True)
    crud.update_user_medical_record_simplified(
        id,
        record_id,
        update_data,
    )
    return Response[object](
        data={},
        message="Historia clínica simplificada actualizada con éxito",
        status_code=200,
        error=None,
    )


@router.patch(
    "/medical_reports/{reporte_id}", response_model=Response[ReporteClinicoResponse]
)
async def update_user_medical_record_simplified(
    id: int,
    record_id: int,
    record: str = Form(...),
    attachments: Optional[List[UploadFile]] = File(None),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[object]:
    """Endpoint para actualizar historias clínicas simplificadas (solo el registro principal)"""
    import json

    try:
        # Parsear el JSON string
        record_data = json.loads(record)

        # Convertir a objeto Pydantic
        record_obj = UpdateUserMedicalRecordRequestDTO(**record_data)

        # Manejar archivos adjuntos
        attachment_urls = []
        if attachments:
            for attachment in attachments:
                if attachment and attachment.filename:
                    # Generar nombre único para el archivo
                    import uuid
                    import os
                    from datetime import datetime

                    file_extension = os.path.splitext(attachment.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"

                    # Crear ruta en S3
                    s3_path = f"medical_records/{id}/{record_id}/{unique_filename}"

                    # Subir archivo a S3
                    try:
                        file_content = await attachment.read()
                        s3_url = crud.upload_file_to_s3(
                            file_content, "images-care-link", s3_path
                        )
                        attachment_urls.append(s3_url)
                    except Exception as e:
                        print(f"Error al subir archivo a S3: {e}")
                        continue

        # Actualizar el record con las URLs de los archivos adjuntos
        update_data = record_obj.dict(exclude_unset=True)
        if attachment_urls:
            update_data["url_hc_adjunto"] = ",".join(attachment_urls)

        crud.update_user_medical_record_simplified(
            id,
            record_id,
            update_data,
        )
        return Response[object](
            data={},
            message="Historia clínica simplificada actualizada con éxito",
            status_code=200,
            error=None,
        )
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Error al parsear JSON: {str(e)}")
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al actualizar historia clínica: {str(e)}"
        )


@router.patch(
    "/medical_reports/{reporte_id}", response_model=Response[ReporteClinicoResponse]
)
def update_reporte_clinico(
    reporte_id: int,
    reporte: ReporteClinicoUpdate,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ReporteClinicoResponse]:
    report_to_update = ReportesClinicos(**reporte.__dict__)
    result = crud.update_medical_record(reporte_id, report_to_update)
    return Response[ReporteClinicoResponse](
        data=ReporteClinicoResponse(**result.__dict__),
        status_code=HTTPStatus.OK,
        message="Reporte clínico actualizado de manera exitosa",
        error=None,
    )


@router.patch(
    "/evolutions/{evolution_id}", response_model=Response[ClinicalEvolutionResponse]
)
def update_evolution(
    evolution_id: int,
    evolution: ClinicalEvolutionUpdate,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ClinicalEvolutionResponse]:
    report_to_update = EvolucionesClinicas(**evolution.__dict__)
    result = crud.update_clinical_evolution(evolution_id, report_to_update)
    return Response[ClinicalEvolutionResponse](
        data=ClinicalEvolutionResponse(**result.__dict__),
        status_code=HTTPStatus.OK,
        message="Evolución clínica actualizada de manera exitosa",
        error=None,
    )


@router.patch("/activities/{id}", response_model=Response[ActivitiesResponse])
def update_activity(
    id: int,
    activity: ActividadesGrupalesUpdate,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ActivitiesResponse]:
    activity_to_update = ActividadesGrupales(**activity.__dict__)
    result = crud.update_activity(id, activity_to_update)
    return Response[ActivitiesResponse](
        data=ActivitiesResponse(**result.__dict__),
        message="Actividad actualizada de manera exitosa",
        status_code=HTTPStatus.OK,
        error=None,
    )


@router.delete("/pagos/{id}", response_model=Response[None])
async def delete_payment(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[None]:
    crud.delete_payment(id)
    return Response[None](
        data=None,
        status_code=HTTPStatus.OK,
        message="Pago eliminado de manera exitosa",
        error=None,
    )


@router.delete("/users/{id}", status_code=200, response_model=Response[object])
async def delete_user(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_user(id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Usuario eliminado de manera exitosa",
        error=None,
    )


@router.delete("/family_members/{id}", status_code=200, response_model=Response[object])
async def delete_family_member(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_family_member(id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Acudiente eliminado de manera exitosa",
        error=None,
    )


@router.delete("/records/{id}", status_code=200, response_model=Response[object])
async def delete_record(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_user_medical_record(id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Historia clínica eliminada de manera exitosa",
        error=None,
    )


@router.delete(
    "/records/{id}/vaccine/{vaccine_id}",
    status_code=200,
    response_model=Response[object],
)
async def delete_vaccine(
    id: int,
    vaccine_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_user_vaccine_by_record_id(id, vaccine_id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Vacuna eliminada de manera exitosa",
        error=None,
    )


@router.delete(
    "/records/{id}/medicine/{medicine_id}",
    status_code=200,
    response_model=Response[object],
)
async def delete_medicine(
    id: int,
    medicine_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_user_medicines_by_record_id(id, medicine_id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Medicamento eliminado de manera exitosa",
        error=None,
    )


@router.delete(
    "/records/{id}/care/{care_id}",
    status_code=200,
    response_model=Response[object],
)
async def delete_care(
    id: int,
    care_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_user_care_by_record_id(id, care_id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Cuidado eliminado de manera exitosa",
        error=None,
    )


@router.delete(
    "/records/{id}/intervention/{intervention_id}",
    status_code=200,
    response_model=Response[object],
)
async def delete_intervention(
    id: int,
    intervention_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_user_intervention_by_record_id(id, intervention_id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Intervención eliminada de manera exitosa",
        error=None,
    )


@router.delete("/evolutions/{id}", status_code=200, response_model=Response[object])
async def delete_evolution(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    crud.delete_clinical_evolution(id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Evolución clínica eliminada de manera exitosa",
        error=None,
    )


@router.delete("/reports/{id}", status_code=200, response_model=Response[object])
async def delete_report(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    crud.delete_medical_report(id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Reporte clínico eliminado de manera exitosa",
        error=None,
    )


@router.delete(
    "/activities/{id}",
    status_code=200,
    response_model=Response[object],
)
async def delete_activity(
    id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[object]:
    crud.delete_activity(id)
    return Response[object](
        data={},
        status_code=HTTPStatus.OK,
        message="Actividad eliminada de manera exitosa",
        error=None,
    )


@router.get("/tarifas-servicio/{id_servicio}/{anio}")
def get_tarifa_servicio(
    id_servicio: int,
    anio: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    """Obtener la tarifa de un servicio para un año específico"""
    try:
        tarifa = crud._get_service_rate(id_servicio, anio)
        return Response[dict](
            data={
                "id_servicio": tarifa.id_servicio,
                "anio": tarifa.anio,
                "precio_por_dia": float(tarifa.precio_por_dia),
            },
            message="Tarifa obtenida exitosamente",
            status_code=HTTPStatus.OK,
            error=None,
        )
    except Exception as e:
        raise HTTPException(
            status_code=404,
            detail=f"Tarifa no encontrada para el servicio {id_servicio} en el año {anio}",
        )


@router.get("/contratos", response_model=Response[List[ContratoResponseDTO]])
def listar_todos_contratos(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[ContratoResponseDTO]]:
    """Obtener todos los contratos del sistema"""
    try:
        contratos = crud.get_all_contracts()
        return Response[List[ContratoResponseDTO]](
            data=contratos,
            status_code=HTTPStatus.OK,
            message="Contratos obtenidos exitosamente",
            error=None,
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al obtener contratos: {str(e)}"
        )


@router.post("/contratos/", response_model=Response[ContratoResponseDTO])
def crear_contrato(
    data: ContratoCreateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[ContratoResponseDTO]:
    try:
        # Usar la función CRUD actualizada que retorna contrato y servicios_por_contrato
        result = crud.create_contract(data)
        contrato = result["contrato"]
        servicios_por_contrato = result["servicios_por_contrato"]

        # Obtener fechas de tiquetera y transporte para cronogramas
        fechas_tiquetera = set()
        fechas_transporte = set()
        servicios_db = []

        # Construir servicios_db para la respuesta usando los datos del CRUD
        for servicio_contratado in servicios_por_contrato:
            # Buscar el servicio original en data.servicios para obtener fechas_servicio
            servicio_original = next(
                s
                for s in data.servicios
                if s.id_servicio == servicio_contratado["id_servicio"]
            )

            # Clasificar fechas para cronogramas
            for f in servicio_original.fechas_servicio:
                if servicio_contratado["id_servicio"] == 1:
                    fechas_tiquetera.add(f.fecha)
                elif servicio_contratado["id_servicio"] == 2:
                    fechas_transporte.add(f.fecha)

            # Para la respuesta
            fechas_dto = [
                FechaServicioDTO(fecha=f.fecha)
                for f in servicio_original.fechas_servicio
            ]
            servicios_db.append(
                ServicioContratoDTO(
                    id_servicio_contratado=servicio_contratado[
                        "id_servicio_contratado"
                    ],
                    id_servicio=servicio_contratado["id_servicio"],
                    fecha=servicio_original.fecha,
                    descripcion=servicio_original.descripcion,
                    precio_por_dia=servicio_original.precio_por_dia,
                    fechas_servicio=fechas_dto,
                )
            )

        # Procesar cronogramas de asistencia y transporte
        id_profesional_default = 1

        # 🔴 VALIDACIÓN: Verificar que no haya doble agendamiento antes de procesar
        fechas_conflicto = []
        for fecha in fechas_tiquetera:
            # Buscar si ya existe un cronograma para esta fecha
            cronograma_existente = (
                crud._CareLinkCrud__carelink_session.query(CronogramaAsistencia)
                .filter(
                    CronogramaAsistencia.fecha == fecha,
                    CronogramaAsistencia.id_profesional == id_profesional_default,
                )
                .first()
            )

            if cronograma_existente:
                # Verificar si el paciente ya está agendado para esta fecha
                paciente_ya_agendado = (
                    crud._CareLinkCrud__carelink_session.query(
                        CronogramaAsistenciaPacientes
                    )
                    .filter(
                        CronogramaAsistenciaPacientes.id_cronograma
                        == cronograma_existente.id_cronograma,
                        CronogramaAsistenciaPacientes.id_usuario == data.id_usuario,
                    )
                    .first()
                )

                if paciente_ya_agendado:
                    fecha_formateada = fecha.strftime("%d/%m/%Y")
                    fechas_conflicto.append(fecha_formateada)
        # Si hay fechas en conflicto, lanzar error con todas las fechas
        if fechas_conflicto:
            fechas_str = ", ".join(fechas_conflicto)
            raise HTTPException(
                status_code=400,
                detail=f"El paciente ya tiene servicios agendados en las siguientes fechas: {fechas_str}. "
                f"No se puede crear un doble agendamiento. "
                f"Por favor, revise la agenda y corrija las fechas antes de continuar.",
            )

        # Si llegamos aquí, no hay conflictos de doble agendamiento
        for fecha in fechas_tiquetera:
            # Crear o buscar cronograma de asistencia
            cronograma_existente = (
                crud._CareLinkCrud__carelink_session.query(CronogramaAsistencia)
                .filter(
                    CronogramaAsistencia.fecha == fecha,
                    CronogramaAsistencia.id_profesional == id_profesional_default,
                )
                .first()
            )
            if not cronograma_existente:
                cronograma_asistencia = CronogramaAsistencia(
                    id_profesional=id_profesional_default,
                    fecha=fecha,
                    comentario=f"Generado automáticamente desde contrato {contrato.id_contrato}",
                )
                crud._CareLinkCrud__carelink_session.add(cronograma_asistencia)
                crud._CareLinkCrud__carelink_session.commit()
                crud._CareLinkCrud__carelink_session.refresh(cronograma_asistencia)
            else:
                cronograma_asistencia = cronograma_existente
            # Determinar si ese día requiere transporte
            requiere_transporte = fecha in fechas_transporte
            # Agregar paciente al cronograma
            paciente_cronograma = CronogramaAsistenciaPacientes(
                id_cronograma=cronograma_asistencia.id_cronograma,
                id_usuario=data.id_usuario,
                id_contrato=contrato.id_contrato,
                estado_asistencia="PENDIENTE",
                requiere_transporte=requiere_transporte,
                observaciones=None,
            )
            crud._CareLinkCrud__carelink_session.add(paciente_cronograma)
            crud._CareLinkCrud__carelink_session.commit()
            crud._CareLinkCrud__carelink_session.refresh(paciente_cronograma)
            # Si requiere transporte, crear cronograma_transporte
            if requiere_transporte:
                cronograma_transporte = CronogramaTransporte(
                    id_cronograma_paciente=paciente_cronograma.id_cronograma_paciente,
                    direccion_recogida="Por definir",
                    direccion_entrega="Por definir",
                    hora_recogida=time(8, 0),
                    hora_entrega=time(17, 0),
                    estado="PENDIENTE",
                    observaciones="Generado automáticamente desde contrato",
                )
                crud._CareLinkCrud__carelink_session.add(cronograma_transporte)
                crud._CareLinkCrud__carelink_session.commit()

        # Construir respuesta con los datos del CRUD
        return Response[ContratoResponseDTO](
            data=ContratoResponseDTO(
                id_contrato=contrato.id_contrato,
                id_usuario=contrato.id_usuario,
                tipo_contrato=contrato.tipo_contrato,
                fecha_inicio=contrato.fecha_inicio,
                fecha_fin=contrato.fecha_fin,
                facturar_contrato=contrato.facturar_contrato,
                estado=contrato.estado,
                servicios=servicios_db,
            ),
            message="Contrato creado de manera exitosa",
            status_code=HTTPStatus.CREATED,
            error=None,
        )

    except HTTPException:
        # Re-lanzar HTTPException sin modificar
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al crear contrato: {str(e)}"
        )


@router.get("/contratos/{contract_id}/factura", response_model=Response[FacturaOut])
async def list_contract_bill(
    contract_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[FacturaOut]:
    bill = crud.get_contract_bill(contract_id)
    bill_response = FacturaOut.from_orm(bill)
    return Response[FacturaOut](
        data=bill_response,
        message="Factura listada",
        error=None,
        status_code=HTTPStatus.OK,
    )


@router.get("/contratos/{id_usuario}", response_model=List[ContratoResponseDTO])
def listar_contratos_por_usuario(
    id_usuario: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    try:
        contratos = db.query(Contratos).filter(Contratos.id_usuario == id_usuario).all()

        if not contratos:
            return []

        result = []
        for contrato in contratos:
            servicios_db = (
                db.query(ServiciosPorContrato)
                .filter(ServiciosPorContrato.id_contrato == contrato.id_contrato)
                .all()
            )
            servicios = []
            for s in servicios_db:
                fechas = (
                    db.query(FechasServicio)
                    .filter(
                        FechasServicio.id_servicio_contratado
                        == s.id_servicio_contratado
                    )
                    .all()
                )
                fechas_dto = [FechaServicioDTO(fecha=f.fecha) for f in fechas]

                servicios.append(
                    ServicioContratoDTO(
                        id_servicio=s.id_servicio,
                        id_servicio_contratado=s.id_servicio_contratado,
                        fecha=s.fecha,
                        descripcion=s.descripcion,
                        precio_por_dia=s.precio_por_dia,
                        fechas_servicio=fechas_dto,
                    )
                )

            result.append(
                ContratoResponseDTO(
                    id_contrato=contrato.id_contrato,
                    id_usuario=id_usuario,
                    tipo_contrato=contrato.tipo_contrato,
                    fecha_inicio=contrato.fecha_inicio,
                    fecha_fin=contrato.fecha_fin,
                    facturar_contrato=contrato.facturar_contrato,
                    estado=contrato.estado,  # <-- Se agrega el campo estado
                    servicios=servicios,
                )
            )

        return result

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al listar contratos: {str(e)}"
        )


@router.get("/contrato/{id_contrato}", response_model=ContratoResponseDTO)
def obtener_contrato(
    id_contrato: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    contrato = db.query(Contratos).filter(Contratos.id_contrato == id_contrato).first()

    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")

    servicios_db = (
        db.query(ServiciosPorContrato)
        .filter(ServiciosPorContrato.id_contrato == contrato.id_contrato)
        .all()
    )
    servicios = []
    for s in servicios_db:
        fechas = (
            db.query(FechasServicio)
            .filter(FechasServicio.id_servicio_contratado == s.id_servicio_contratado)
            .all()
        )
        fechas_dto = [FechaServicioDTO(fecha=f.fecha) for f in fechas]

        servicios.append(
            ServicioContratoDTO(
                id_servicio_contratado=s.id_servicio_contratado,
                id_servicio=s.id_servicio,
                fecha=s.fecha,
                descripcion=s.descripcion,
                precio_por_dia=s.precio_por_dia,
                fechas_servicio=fechas_dto,
            )
        )

    return ContratoResponseDTO(
        id_contrato=contrato.id_contrato,
        id_usuario=contrato.id_usuario,
        tipo_contrato=contrato.tipo_contrato,
        fecha_inicio=contrato.fecha_inicio,
        fecha_fin=contrato.fecha_fin,
        facturar_contrato=contrato.facturar_contrato,
        estado=contrato.estado,  # <-- Se agrega el campo estado
        servicios=servicios,
    )


@router.patch("/contrato/{id_contrato}")
def actualizar_contrato(
    id_contrato: int,
    data: ContratoUpdateDTO,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    contrato = db.query(Contratos).filter(Contratos.id_contrato == id_contrato).first()

    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato no encontrado")

    try:
        # 🔴 Eliminar servicios anteriores del contrato (y sus fechas asociadas)
        servicios_anteriores = (
            db.query(ServiciosPorContrato)
            .filter(ServiciosPorContrato.id_contrato == contrato.id_contrato)
            .all()
        )

        for servicio_ant in servicios_anteriores:
            # Eliminar fechas asociadas
            db.query(FechasServicio).filter(
                FechasServicio.id_servicio_contratado
                == servicio_ant.id_servicio_contratado
            ).delete()

        # Eliminar servicios contratados
        db.query(ServiciosPorContrato).filter(
            ServiciosPorContrato.id_contrato == contrato.id_contrato
        ).delete()

        db.commit()

        # 🟢 Agregar nuevos servicios y fechas desde data.servicios
        for servicio in data.servicios:
            servicio_contratado = ServiciosPorContrato(
                id_contrato=contrato.id_contrato,
                id_servicio=servicio.id_servicio,
                fecha=servicio.fecha,
                descripcion=servicio.descripcion,
                precio_por_dia=servicio.precio_por_dia,
            )
            db.add(servicio_contratado)
            db.commit()
            db.refresh(servicio_contratado)

            for f in servicio.fechas_servicio:
                fecha_servicio = FechasServicio(
                    id_servicio_contratado=servicio_contratado.id_servicio_contratado,
                    fecha=f.fecha,
                )
                db.add(fecha_servicio)

        # ✏️ Actualizar otros atributos del contrato
        for attr, value in data.dict(exclude={"servicios"}, exclude_unset=True).items():
            setattr(contrato, attr, value)

        db.commit()
        db.refresh(contrato)

        return {"message": "Contrato actualizado correctamente"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Error al actualizar contrato: {str(e)}"
        )


@router.post("/servicios/{id_servicio}/fechas")
def crear_fechas_servicio(
    id_servicio: int,
    fechas: List[FechaServicioDTO],
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    for fecha in fechas:
        nueva_fecha = FechasServicio(
            id_servicio_contratado=id_servicio, fecha=fecha.fecha
        )
        db.add(nueva_fecha)
    db.commit()

    return {"message": "Fechas creadas correctamente"}


@router.patch("/servicios/{id_servicio_contratado}/fechas")
def actualizar_fechas_servicio(
    id_servicio_contratado: int,
    fechas: List[FechaServicioDTO],
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    servicio = (
        db.query(ServiciosPorContrato)
        .filter_by(id_servicio_contratado=id_servicio_contratado)
        .first()
    )

    if not servicio:
        raise HTTPException(status_code=404, detail="Servicio contratado no encontrado")

    db.query(FechasServicio).filter_by(
        id_servicio_contratado=id_servicio_contratado
    ).delete()

    for f in fechas:
        nueva_fecha = FechasServicio(
            id_servicio_contratado=id_servicio_contratado, fecha=f.fecha
        )
        db.add(nueva_fecha)

    db.commit()
    return {"message": "Fechas actualizadas correctamente"}


@router.post("/pagos/")
def crear_pago(
    data: PagoCreateDTO,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    factura = db.query(Facturas).filter(Facturas.id_factura == data.id_factura).first()
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    total_pagado = (
        db.query(func.sum(Pagos.valor))
        .filter(Pagos.id_factura == data.id_factura)
        .scalar()
        or 0
    )

    if total_pagado + data.valor > factura.total_factura:
        raise HTTPException(
            status_code=400,
            detail=f"El pago excede el total restante de la factura. Total restante: {factura.total_factura - total_pagado:.2f}",
        )

    nuevo_pago = Pagos(
        id_factura=data.id_factura,
        id_metodo_pago=data.id_metodo_pago,
        id_tipo_pago=data.id_tipo_pago,
        fecha_pago=data.fecha_pago,
        valor=data.valor,
    )
    db.add(nuevo_pago)
    db.commit()
    db.refresh(nuevo_pago)
    return {"message": "Pago registrado correctamente", "id_pago": nuevo_pago.id_pago}


@router.get("/pagos/factura/{factura_id}", response_model=list[PagoResponseDTO])
def get_pagos_by_factura(
    factura_id: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    pagos = db.query(Pagos).filter(Pagos.id_factura == factura_id).all()
    return [PagoResponseDTO.from_orm(pago) for pago in pagos]


@router.delete("/pagos/{id_pago}")
def eliminar_pago(
    id_pago: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    pago = db.query(Pagos).filter(Pagos.id_pago == id_pago).first()
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    db.delete(pago)
    db.commit()
    return {"message": "Pago eliminado correctamente"}


# ============================================================================
# ENDPOINTS DE FACTURACIÓN - IMPLEMENTACIÓN COMPLETA
# ============================================================================


# ============================================================================
# ENDPOINTS DE CRONOGRAMA DE ASISTENCIA
# ============================================================================


@router.post(
    "/cronograma_asistencia/crear",
    response_model=Response[CronogramaAsistenciaResponseDTO],
)
def crear_cronograma_asistencia(
    cronograma_data: CronogramaAsistenciaCreateDTO,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[CronogramaAsistenciaResponseDTO]:
    """
    Crea un nuevo cronograma de asistencia
    """
    try:
        # Verificar si ya existe un cronograma para esta fecha y profesional
        cronograma_existente = (
            db.query(CronogramaAsistencia)
            .filter(
                CronogramaAsistencia.fecha == cronograma_data.fecha,
                CronogramaAsistencia.id_profesional == cronograma_data.id_profesional,
            )
            .first()
        )

        if cronograma_existente:
            # Si ya existe, retornar el existente
            return Response[CronogramaAsistenciaResponseDTO](
                data=CronogramaAsistenciaResponseDTO(
                    id_cronograma=cronograma_existente.id_cronograma,
                    id_profesional=cronograma_existente.id_profesional,
                    fecha=cronograma_existente.fecha,
                    comentario=cronograma_existente.comentario,
                    fecha_creacion=cronograma_existente.fecha_creacion,
                    fecha_actualizacion=cronograma_existente.fecha_actualizacion,
                    pacientes=[],
                ),
                status_code=HTTPStatus.OK,
                message="Cronograma existente recuperado",
                error=None,
            )

        # Crear nuevo cronograma
        nuevo_cronograma = CronogramaAsistencia(
            id_profesional=cronograma_data.id_profesional,
            fecha=cronograma_data.fecha,
            comentario=cronograma_data.comentario,
        )

        db.add(nuevo_cronograma)
        db.commit()
        db.refresh(nuevo_cronograma)

        return Response[CronogramaAsistenciaResponseDTO](
            data=CronogramaAsistenciaResponseDTO(
                id_cronograma=nuevo_cronograma.id_cronograma,
                id_profesional=nuevo_cronograma.id_profesional,
                fecha=nuevo_cronograma.fecha,
                comentario=nuevo_cronograma.comentario,
                fecha_creacion=nuevo_cronograma.fecha_creacion,
                fecha_actualizacion=nuevo_cronograma.fecha_actualizacion,
                pacientes=[],
            ),
            status_code=HTTPStatus.CREATED,
            message="Cronograma creado exitosamente",
            error=None,
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Error al crear cronograma: {str(e)}"
        )


@router.post(
    "/cronograma_asistencia/paciente/agregar",
    response_model=Response[CronogramaAsistenciaPacienteResponseDTO],
)
def agregar_paciente_cronograma(
    paciente_data: CronogramaAsistenciaPacienteCreateDTO,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[CronogramaAsistenciaPacienteResponseDTO]:
    """
    Agrega un paciente a un cronograma de asistencia existente
    """
    try:
        # 🔴 VALIDACIÓN: Verificar que el cronograma existe
        cronograma = (
            db.query(CronogramaAsistencia)
            .filter(CronogramaAsistencia.id_cronograma == paciente_data.id_cronograma)
            .first()
        )

        if not cronograma:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el cronograma con ID {paciente_data.id_cronograma}",
            )

        # 🔴 VALIDACIÓN: Verificar que el paciente ya está en el cronograma
        paciente_existente = (
            db.query(CronogramaAsistenciaPacientes)
            .filter(
                CronogramaAsistenciaPacientes.id_cronograma
                == paciente_data.id_cronograma,
                CronogramaAsistenciaPacientes.id_usuario == paciente_data.id_usuario,
            )
            .first()
        )

        if paciente_existente:
            raise HTTPException(
                status_code=400,
                detail=f"El paciente ya está registrado en este cronograma para la fecha {cronograma.fecha}. "
                f"No se puede crear un doble agendamiento. "
                f"Paciente ID: {paciente_data.id_usuario}, "
                f"Estado actual: {paciente_existente.estado_asistencia}",
            )

        # 🔴 VALIDACIÓN: Verificar que el contrato existe y pertenece al usuario
        contrato = (
            db.query(Contratos)
            .filter(
                Contratos.id_contrato == paciente_data.id_contrato,
                Contratos.id_usuario == paciente_data.id_usuario,
            )
            .first()
        )

        if not contrato:
            raise HTTPException(
                status_code=400,
                detail=f"El contrato {paciente_data.id_contrato} no existe o no pertenece al usuario {paciente_data.id_usuario}",
            )

        # Agregar paciente al cronograma
        nuevo_paciente = CronogramaAsistenciaPacientes(
            id_cronograma=paciente_data.id_cronograma,
            id_usuario=paciente_data.id_usuario,
            id_contrato=paciente_data.id_contrato,
            estado_asistencia=paciente_data.estado_asistencia,
            observaciones=paciente_data.observaciones,
        )

        db.add(nuevo_paciente)
        db.commit()
        db.refresh(nuevo_paciente)

        return Response[CronogramaAsistenciaPacienteResponseDTO](
            data=CronogramaAsistenciaPacienteResponseDTO(
                id_cronograma_paciente=nuevo_paciente.id_cronograma_paciente,
                id_cronograma=nuevo_paciente.id_cronograma,
                id_usuario=nuevo_paciente.id_usuario,
                id_contrato=nuevo_paciente.id_contrato,
                estado_asistencia=nuevo_paciente.estado_asistencia,
                requiere_transporte=nuevo_paciente.requiere_transporte,
                observaciones=nuevo_paciente.observaciones,
            ),
            status_code=HTTPStatus.CREATED,
            message="Paciente agregado al cronograma exitosamente",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Error al agregar paciente al cronograma: {str(e)}"
        )


@router.get(
    "/cronograma_asistencia/rango/{fecha_inicio}/{fecha_fin}",
    response_model=Response[List[CronogramaAsistenciaResponseDTO]],
)
def get_cronogramas_por_rango(
    fecha_inicio: str,
    fecha_fin: str,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[CronogramaAsistenciaResponseDTO]]:
    """
    Obtiene los cronogramas de asistencia en un rango de fechas
    """
    try:
        # Convertir fechas de string a date
        fecha_inicio_date = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin_date = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

        # Consultar cronogramas en el rango
        cronogramas = (
            db.query(CronogramaAsistencia)
            .filter(
                CronogramaAsistencia.fecha >= fecha_inicio_date,
                CronogramaAsistencia.fecha <= fecha_fin_date,
            )
            .all()
        )

        result = []
        for cronograma in cronogramas:
            # Obtener pacientes agendados para este cronograma
            pacientes_agendados = (
                db.query(CronogramaAsistenciaPacientes, User)
                .join(User, CronogramaAsistenciaPacientes.id_usuario == User.id_usuario)
                .filter(
                    CronogramaAsistenciaPacientes.id_cronograma
                    == cronograma.id_cronograma
                )
                .all()
            )

            pacientes_dto = []
            for paciente_agendado, usuario in pacientes_agendados:
                # Obtener información de transporte si existe
                transporte_info = None
                if paciente_agendado.requiere_transporte:
                    transporte = (
                        db.query(CronogramaTransporte)
                        .filter(
                            CronogramaTransporte.id_cronograma_paciente
                            == paciente_agendado.id_cronograma_paciente
                        )
                        .first()
                    )
                    if transporte:
                        transporte_info = {
                            "id_transporte": transporte.id_transporte,
                            "direccion_recogida": transporte.direccion_recogida,
                            "direccion_entrega": transporte.direccion_entrega,
                            "hora_recogida": (
                                str(transporte.hora_recogida)
                                if transporte.hora_recogida
                                else None
                            ),
                            "hora_entrega": (
                                str(transporte.hora_entrega)
                                if transporte.hora_entrega
                                else None
                            ),
                            "estado": transporte.estado,
                            "observaciones": transporte.observaciones,
                        }
                pacientes_dto.append(
                    PacientePorFechaDTO(
                        id_cronograma_paciente=paciente_agendado.id_cronograma_paciente,
                        id_usuario=paciente_agendado.id_usuario,
                        id_contrato=paciente_agendado.id_contrato,
                        estado_asistencia=paciente_agendado.estado_asistencia,
                        requiere_transporte=paciente_agendado.requiere_transporte,
                        nombres=usuario.nombres,
                        apellidos=usuario.apellidos,
                        n_documento=usuario.n_documento,
                        transporte_info=transporte_info,
                        fecha_creacion=paciente_agendado.fecha_creacion,
                        fecha_actualizacion=paciente_agendado.fecha_actualizacion,
                    )
                )

            result.append(
                CronogramaAsistenciaResponseDTO(
                    id_cronograma=cronograma.id_cronograma,
                    id_profesional=cronograma.id_profesional,
                    fecha=cronograma.fecha,
                    comentario=cronograma.comentario,
                    fecha_creacion=cronograma.fecha_creacion,
                    fecha_actualizacion=cronograma.fecha_actualizacion,
                    pacientes=pacientes_dto,
                )
            )

        return Response[List[CronogramaAsistenciaResponseDTO]](
            data=result,
            status_code=HTTPStatus.OK,
            message="Cronogramas consultados exitosamente",
            error=None,
        )

    except ValueError as e:
        raise HTTPException(
            status_code=400, detail=f"Formato de fecha inválido: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.get(
    "/cronograma_asistencia/profesional/{id_profesional}",
    response_model=Response[List[CronogramaAsistenciaResponseDTO]],
)
def get_cronogramas_por_profesional(
    id_profesional: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[CronogramaAsistenciaResponseDTO]]:
    """
    Obtiene los cronogramas de asistencia de un profesional específico
    """
    try:
        # Consultar cronogramas del profesional
        cronogramas = (
            db.query(CronogramaAsistencia)
            .filter(CronogramaAsistencia.id_profesional == id_profesional)
            .all()
        )

        result = []
        for cronograma in cronogramas:
            # Obtener pacientes agendados para este cronograma
            pacientes_agendados = (
                db.query(CronogramaAsistenciaPacientes, User)
                .join(User, CronogramaAsistenciaPacientes.id_usuario == User.id_usuario)
                .filter(
                    CronogramaAsistenciaPacientes.id_cronograma
                    == cronograma.id_cronograma
                )
                .all()
            )

            pacientes_dto = []
            for paciente_agendado, usuario in pacientes_agendados:
                # Obtener información de transporte si existe
                transporte_info = None
                if paciente_agendado.requiere_transporte:
                    transporte = (
                        db.query(CronogramaTransporte)
                        .filter(
                            CronogramaTransporte.id_cronograma_paciente
                            == paciente_agendado.id_cronograma_paciente
                        )
                        .first()
                    )
                    if transporte:
                        transporte_info = {
                            "id_transporte": transporte.id_transporte,
                            "direccion_recogida": transporte.direccion_recogida,
                            "direccion_entrega": transporte.direccion_entrega,
                            "hora_recogida": (
                                str(transporte.hora_recogida)
                                if transporte.hora_recogida
                                else None
                            ),
                            "hora_entrega": (
                                str(transporte.hora_entrega)
                                if transporte.hora_entrega
                                else None
                            ),
                            "estado": transporte.estado,
                            "observaciones": transporte.observaciones,
                        }
                pacientes_dto.append(
                    PacientePorFechaDTO(
                        id_cronograma_paciente=paciente_agendado.id_cronograma_paciente,
                        id_usuario=paciente_agendado.id_usuario,
                        id_contrato=paciente_agendado.id_contrato,
                        estado_asistencia=paciente_agendado.estado_asistencia,
                        requiere_transporte=paciente_agendado.requiere_transporte,
                        nombres=usuario.nombres,
                        apellidos=usuario.apellidos,
                        n_documento=usuario.n_documento,
                        transporte_info=transporte_info,
                        fecha_creacion=paciente_agendado.fecha_creacion,
                        fecha_actualizacion=paciente_agendado.fecha_actualizacion,
                    )
                )

            result.append(
                CronogramaAsistenciaResponseDTO(
                    id_cronograma=cronograma.id_cronograma,
                    id_profesional=cronograma.id_profesional,
                    fecha=cronograma.fecha,
                    comentario=cronograma.comentario,
                    fecha_creacion=cronograma.fecha_creacion,
                    fecha_actualizacion=cronograma.fecha_actualizacion,
                    pacientes=pacientes_dto,
                )
            )

        return Response[List[CronogramaAsistenciaResponseDTO]](
            data=result,
            status_code=HTTPStatus.OK,
            message="Cronogramas del profesional consultados exitosamente",
            error=None,
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.patch(
    "/cronograma_asistencia/paciente/{id_cronograma_paciente}/estado",
    response_model=Response[CronogramaAsistenciaPacienteResponseDTO],
)
def update_estado_asistencia(
    id_cronograma_paciente: int,
    estado_data: EstadoAsistenciaUpdateDTO,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[CronogramaAsistenciaPacienteResponseDTO]:
    """
    Actualiza el estado de asistencia de un paciente y guarda observaciones.
    Si asiste o no asiste (sin justificación), descuenta día de tiquetera.
    Si se agotan días, marca contrato como vencido y genera alerta.
    """
    try:
        paciente_cronograma = (
            db.query(CronogramaAsistenciaPacientes)
            .filter(
                CronogramaAsistenciaPacientes.id_cronograma_paciente
                == id_cronograma_paciente
            )
            .first()
        )
        if not paciente_cronograma:
            raise HTTPException(
                status_code=404, detail="Paciente no encontrado en el cronograma"
            )

        # 🔴 VALIDACIÓN: Solo se puede cambiar estado de registros "PENDIENTE"
        if paciente_cronograma.estado_asistencia != "PENDIENTE":
            raise HTTPException(
                status_code=400,
                detail=f"No se puede cambiar el estado de un paciente con estado '{paciente_cronograma.estado_asistencia}'. Solo se puede modificar registros con estado 'PENDIENTE'.",
            )

        estados_validos = ["PENDIENTE", "ASISTIO", "NO_ASISTIO", "CANCELADO"]
        if estado_data.estado_asistencia not in estados_validos:
            raise HTTPException(
                status_code=400,
                detail=f"Estado inválido. Estados válidos: {', '.join(estados_validos)}",
            )
        # Guardar observaciones
        paciente_cronograma.observaciones = estado_data.observaciones
        paciente_cronograma.estado_asistencia = estado_data.estado_asistencia
        db.commit()
        db.refresh(paciente_cronograma)

        # 🔴 LÓGICA DE DESCUENTO DE DÍAS DE TIQUETERA
        # Se descuenta día tanto si ASISTE como si NO ASISTE (sin justificación)
        if estado_data.estado_asistencia in ["ASISTIO", "NO_ASISTIO"]:
            contrato = (
                db.query(Contratos)
                .filter(Contratos.id_contrato == paciente_cronograma.id_contrato)
                .first()
            )
            if contrato:
                # Contar días consumidos (asistencias + no asistencias)
                total_dias_consumidos = (
                    db.query(CronogramaAsistenciaPacientes)
                    .filter(
                        CronogramaAsistenciaPacientes.id_contrato
                        == contrato.id_contrato,
                        CronogramaAsistenciaPacientes.estado_asistencia.in_(
                            ["ASISTIO", "NO_ASISTIO"]
                        ),
                    )
                    .count()
                )

                # Contar total de días de la tiquetera (servicio 1)
                total_tiquetera = (
                    db.query(CronogramaAsistenciaPacientes)
                    .filter(
                        CronogramaAsistenciaPacientes.id_contrato
                        == contrato.id_contrato
                    )
                    .count()
                )

                # Si se agotaron todos los días, marcar contrato como vencido
                if total_dias_consumidos >= total_tiquetera:
                    contrato.estado = "VENCIDO"
                    db.commit()
                    # Aquí puedes agregar lógica para generar una alerta al profesional
        response_dto = CronogramaAsistenciaPacienteResponseDTO(
            id_cronograma_paciente=paciente_cronograma.id_cronograma_paciente,
            id_cronograma=paciente_cronograma.id_cronograma,
            id_usuario=paciente_cronograma.id_usuario,
            id_contrato=paciente_cronograma.id_contrato,
            estado_asistencia=paciente_cronograma.estado_asistencia,
            requiere_transporte=paciente_cronograma.requiere_transporte,
            observaciones=paciente_cronograma.observaciones,
            fecha_creacion=paciente_cronograma.fecha_creacion,
            fecha_actualizacion=paciente_cronograma.fecha_actualizacion,
        )
        return Response[CronogramaAsistenciaPacienteResponseDTO](
            data=response_dto,
            status_code=HTTPStatus.OK,
            message="Estado de asistencia actualizado exitosamente",
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.post(
    "/cronograma_asistencia/paciente/{id_cronograma_paciente}/reagendar",
    response_model=Response[CronogramaAsistenciaPacienteResponseDTO],
)
def reagendar_asistencia_paciente(
    id_cronograma_paciente: int,
    estado_data: EstadoAsistenciaUpdateDTO,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[CronogramaAsistenciaPacienteResponseDTO]:
    """
    Reagenda la asistencia de un paciente SOLO si existe justificación (observaciones) y nueva fecha.
    """
    try:
        paciente_cronograma = (
            db.query(CronogramaAsistenciaPacientes)
            .filter(
                CronogramaAsistenciaPacientes.id_cronograma_paciente
                == id_cronograma_paciente
            )
            .first()
        )
        if not paciente_cronograma:
            raise HTTPException(
                status_code=404, detail="Paciente no encontrado en el cronograma"
            )

        # 🔴 VALIDACIÓN POR REGISTRO INDIVIDUAL Y ESTADO
        # Solo se puede reagendar si el estado es "PENDIENTE"
        if paciente_cronograma.estado_asistencia in [
            "REAGENDADO",
            "ASISTIO",
            "NO_ASISTIO",
            "CANCELADO",
        ]:
            raise HTTPException(
                status_code=400,
                detail=f"No se puede reagendar un paciente que ya tiene estado '{paciente_cronograma.estado_asistencia}' en esta fecha. Solo se puede reagendar si el estado es 'PENDIENTE'.",
            )

        if not estado_data.observaciones or not estado_data.observaciones.strip():
            raise HTTPException(
                status_code=400,
                detail="No se puede reagendar sin justificación en observaciones.",
            )
        if not estado_data.nueva_fecha:
            raise HTTPException(
                status_code=400,
                detail="Debe seleccionar una nueva fecha para reagendar.",
            )

        # 🔴 PASO 1: Obtener el id_profesional del cronograma original
        cronograma_original = (
            db.query(CronogramaAsistencia)
            .filter_by(id_cronograma=paciente_cronograma.id_cronograma)
            .first()
        )
        id_profesional = cronograma_original.id_profesional

        # 🔴 PASO 2: Buscar o crear cronograma para la nueva fecha
        cronograma_nuevo = (
            db.query(CronogramaAsistencia)
            .filter_by(id_profesional=id_profesional, fecha=estado_data.nueva_fecha)
            .first()
        )

        if not cronograma_nuevo:
            cronograma_nuevo = CronogramaAsistencia(
                id_profesional=id_profesional, fecha=estado_data.nueva_fecha
            )
            db.add(cronograma_nuevo)
            db.commit()
            db.refresh(cronograma_nuevo)

        # 🔴 PASO 3: Verificar que no esté ya agendado para la nueva fecha
        agendado_existente = (
            db.query(CronogramaAsistenciaPacientes)
            .filter_by(
                id_cronograma=cronograma_nuevo.id_cronograma,
                id_usuario=paciente_cronograma.id_usuario,
            )
            .first()
        )

        if agendado_existente:
            raise HTTPException(
                status_code=400, detail="El paciente ya está agendado para esta fecha."
            )

        # 🔴 PASO 4: Crear nuevo registro para la nueva fecha
        nuevo_paciente_cronograma = CronogramaAsistenciaPacientes(
            id_cronograma=cronograma_nuevo.id_cronograma,
            id_usuario=paciente_cronograma.id_usuario,
            id_contrato=paciente_cronograma.id_contrato,
            estado_asistencia="PENDIENTE",
            observaciones="",
        )
        db.add(nuevo_paciente_cronograma)
        db.commit()
        db.refresh(nuevo_paciente_cronograma)

        # 🔴 PASO 5: SOLO DESPUÉS de crear exitosamente el nuevo registro, cambiar estado del original
        paciente_cronograma.estado_asistencia = "REAGENDADO"
        paciente_cronograma.observaciones = estado_data.observaciones
        db.commit()

        response_dto = CronogramaAsistenciaPacienteResponseDTO(
            id_cronograma_paciente=nuevo_paciente_cronograma.id_cronograma_paciente,
            id_cronograma=nuevo_paciente_cronograma.id_cronograma,
            id_usuario=nuevo_paciente_cronograma.id_usuario,
            id_contrato=nuevo_paciente_cronograma.id_contrato,
            estado_asistencia=nuevo_paciente_cronograma.estado_asistencia,
            requiere_transporte=nuevo_paciente_cronograma.requiere_transporte,
            observaciones=nuevo_paciente_cronograma.observaciones,
            fecha_creacion=nuevo_paciente_cronograma.fecha_creacion,
            fecha_actualizacion=nuevo_paciente_cronograma.fecha_actualizacion,
        )
        return Response[CronogramaAsistenciaPacienteResponseDTO](
            data=response_dto,
            status_code=HTTPStatus.OK,
            message="Paciente reagendado exitosamente",
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


# ============================================================================
# ENDPOINTS DE TRANSPORTE MOVIDOS A transporte_controller.py
# ============================================================================
# Todos los endpoints de transporte han sido movidos al controlador específico
# para evitar conflictos y mantener una mejor organización del código

# ============================================================================
# ENDPOINTS DE FACTURACIÓN - IMPLEMENTACIÓN COMPLETA
# ============================================================================


@router.post("/facturas/", response_model=Response[FacturaOut])
def create_factura(
    factura_data: FacturaCreate,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[FacturaOut]:
    """
    Crea una nueva factura con sus pagos asociados

    Args:
        factura_data: Datos de la factura incluyendo pagos
        db: Sesión de base de datos
        _: Usuario autenticado

    Returns:
        Response con la factura creada

    Raises:
        HTTPException: Si el contrato no existe o hay errores de validación
    """
    try:
        # Validar que el contrato existe
        contrato = (
            db.query(Contratos)
            .filter(Contratos.id_contrato == factura_data.id_contrato)
            .first()
        )

        if not contrato:
            raise HTTPException(
                status_code=404,
                detail=f"Contrato con ID {factura_data.id_contrato} no encontrado",
            )

        # Validar que el contrato esté activo
        if contrato.estado != "ACTIVO":
            raise HTTPException(
                status_code=400,
                detail=f"El contrato {contrato.id_contrato} no está activo (estado: {contrato.estado})",
            )

        # Crear la factura
        factura = Facturas(
            id_contrato=factura_data.id_contrato,
            fecha_emision=factura_data.fecha_emision,
            fecha_vencimiento=factura_data.fecha_vencimiento,
            total_factura=factura_data.total,  # Corregido: usar total_factura
            subtotal=factura_data.total,  # Por defecto igual al total
            impuestos=0.00,
            descuentos=0.00,
            estado_factura=EstadoFactura.PENDIENTE,
        )

        db.add(factura)
        db.flush()  # Para obtener el ID de la factura

        # Procesar pagos si existen
        total_pagos = 0
        for pago_data in factura_data.pagos:
            # Validar método de pago
            metodo_pago = (
                db.query(MetodoPago)
                .filter(MetodoPago.id_metodo_pago == pago_data.id_metodo_pago)
                .first()
            )

            if not metodo_pago:
                raise HTTPException(
                    status_code=400,
                    detail=f"Método de pago con ID {pago_data.id_metodo_pago} no existe",
                )

            # Validar tipo de pago
            tipo_pago = (
                db.query(TipoPago)
                .filter(TipoPago.id_tipo_pago == pago_data.id_tipo_pago)
                .first()
            )

            if not tipo_pago:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tipo de pago con ID {pago_data.id_tipo_pago} no existe",
                )

            # Crear el pago
            pago = Pagos(
                id_factura=factura.id_factura,
                id_metodo_pago=pago_data.id_metodo_pago,
                id_tipo_pago=pago_data.id_tipo_pago,
                fecha_pago=pago_data.fecha_pago,
                valor=pago_data.valor,
            )

            db.add(pago)
            total_pagos += float(pago_data.valor)

        # Validar que los pagos no excedan el total de la factura
        if total_pagos > float(factura_data.total):
            raise HTTPException(
                status_code=400,
                detail=f"El total de pagos ({total_pagos}) excede el total de la factura ({factura_data.total})",
            )

        # Actualizar estado de la factura si los pagos cubren el total
        if total_pagos >= float(factura_data.total):
            factura.estado_factura = EstadoFactura.PAGADA

        db.commit()
        db.refresh(factura)

        # Crear respuesta
        factura_response = FacturaOut(
            id_factura=factura.id_factura,
            numero_factura=factura.numero_factura,
            id_contrato=factura.id_contrato,
            fecha_emision=factura.fecha_emision,
            total_factura=float(factura.total_factura),
        )

        return Response[FacturaOut](
            data=factura_response,
            status_code=HTTPStatus.CREATED,
            message="Factura creada exitosamente",
            error=None,
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor al crear factura: {str(e)}",
        )


@router.post("/facturas/{factura_id}/pagos/", response_model=Response[dict])
def add_pago_to_factura(
    factura_id: int,
    pagos: List[PagoCreate],
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[dict]:
    """
    Agrega uno o varios pagos a una factura existente

    Args:
        factura_id: ID de la factura
        pagos: Lista de pagos a agregar
        db: Sesión de base de datos
        _: Usuario autenticado

    Returns:
        Response con mensaje de confirmación

    Raises:
        HTTPException: Si la factura no existe o hay errores de validación
    """
    try:
        # Validar que la factura existe
        factura = db.query(Facturas).filter(Facturas.id_factura == factura_id).first()

        if not factura:
            raise HTTPException(
                status_code=404, detail=f"Factura con ID {factura_id} no encontrada"
            )

        # Validar que la factura no esté cancelada o anulada
        if factura.estado_factura in [EstadoFactura.CANCELADA, EstadoFactura.ANULADA]:
            raise HTTPException(
                status_code=400,
                detail=f"No se pueden agregar pagos a una factura con estado {factura.estado_factura.value}",
            )

        # Calcular total de pagos existentes
        total_existente = sum(float(p.valor) for p in factura.pagos)
        total_nuevo = sum(float(p.valor) for p in pagos)
        total_factura = float(factura.total_factura)

        # Validar que los nuevos pagos no excedan el total
        if total_existente + total_nuevo > total_factura:
            raise HTTPException(
                status_code=400,
                detail=f"Los pagos exceden el total de la factura. "
                f"Total factura: {total_factura}, "
                f"Pagos existentes: {total_existente}, "
                f"Nuevos pagos: {total_nuevo}",
            )

        # Procesar cada pago
        for pago_data in pagos:
            # Validar método de pago
            metodo_pago = (
                db.query(MetodoPago)
                .filter(MetodoPago.id_metodo_pago == pago_data.id_metodo_pago)
                .first()
            )

            if not metodo_pago:
                raise HTTPException(
                    status_code=400,
                    detail=f"Método de pago con ID {pago_data.id_metodo_pago} no existe",
                )

            # Validar tipo de pago
            tipo_pago = (
                db.query(TipoPago)
                .filter(TipoPago.id_tipo_pago == pago_data.id_tipo_pago)
                .first()
            )

            if not tipo_pago:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tipo de pago con ID {pago_data.id_tipo_pago} no existe",
                )

            # Crear el pago
            pago = Pagos(
                id_factura=factura_id,
                id_metodo_pago=pago_data.id_metodo_pago,
                id_tipo_pago=pago_data.id_tipo_pago,
                fecha_pago=pago_data.fecha_pago,
                valor=pago_data.valor,
            )

            db.add(pago)

        # Actualizar estado de la factura si los pagos cubren el total
        if total_existente + total_nuevo >= total_factura:
            factura.estado_factura = EstadoFactura.PAGADA

        db.commit()

        return Response[dict](
            data={"factura_id": factura_id, "pagos_agregados": len(pagos)},
            status_code=HTTPStatus.OK,
            message=f"Se agregaron {len(pagos)} pagos a la factura exitosamente",
            error=None,
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor al agregar pagos: {str(e)}",
        )


@router.delete("/facturas/{factura_id}", response_model=Response[dict])
def delete_factura(
    factura_id: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[dict]:
    """
    Elimina una factura y todos sus pagos asociados

    Args:
        factura_id: ID de la factura a eliminar
        db: Sesión de base de datos
        _: Usuario autenticado

    Returns:
        Response con mensaje de confirmación

    Raises:
        HTTPException: Si la factura no existe o no se puede eliminar
    """
    try:
        # Validar que la factura existe
        factura = db.query(Facturas).filter(Facturas.id_factura == factura_id).first()

        if not factura:
            raise HTTPException(
                status_code=404, detail=f"Factura con ID {factura_id} no encontrada"
            )

        # Validar que la factura no esté pagada (opcional, según política de negocio)
        if factura.estado_factura == EstadoFactura.PAGADA:
            raise HTTPException(
                status_code=400,
                detail="No se puede eliminar una factura que ya ha sido pagada",
            )

        # Eliminar la factura (los pagos se eliminan automáticamente por CASCADE)
        db.delete(factura)
        db.commit()

        return Response[dict](
            data={"factura_id": factura_id},
            status_code=HTTPStatus.OK,
            message="Factura eliminada exitosamente",
            error=None,
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor al eliminar factura: {str(e)}",
        )


@router.patch("/facturas/{factura_id}", response_model=Response[FacturaOut])
def update_factura(
    factura_id: int,
    factura_data: dict,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[FacturaOut]:
    """
    Actualiza una factura existente

    Args:
        factura_id: ID de la factura a actualizar
        factura_data: Datos a actualizar
        db: Sesión de base de datos
        _: Usuario autenticado

    Returns:
        Response con la factura actualizada

    Raises:
        HTTPException: Si la factura no existe o hay errores de validación
    """
    try:
        # Validar que la factura existe
        factura = db.query(Facturas).filter(Facturas.id_factura == factura_id).first()

        if not factura:
            raise HTTPException(
                status_code=404, detail=f"Factura con ID {factura_id} no encontrada"
            )

        # Validar que la factura no esté pagada (para ciertos campos)
        campos_restringidos = ["total_factura", "subtotal", "impuestos", "descuentos"]
        if factura.estado_factura == EstadoFactura.PAGADA:
            for campo in campos_restringidos:
                if campo in factura_data:
                    raise HTTPException(
                        status_code=400,
                        detail=f"No se puede modificar el campo '{campo}' en una factura pagada",
                    )

        # Actualizar campos permitidos
        campos_permitidos = {
            "fecha_emision": "fecha_emision",
            "fecha_vencimiento": "fecha_vencimiento",
            "total_factura": "total_factura",
            "subtotal": "subtotal",
            "impuestos": "impuestos",
            "descuentos": "descuentos",
            "estado_factura": "estado_factura",
            "observaciones": "observaciones",
        }

        for campo_request, campo_modelo in campos_permitidos.items():
            if campo_request in factura_data:
                valor = factura_data[campo_request]

                # Validaciones específicas
                if campo_request == "estado_factura":
                    try:
                        valor = EstadoFactura(valor)
                    except ValueError:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Estado de factura inválido: {valor}. "
                            f"Estados válidos: {[e.value for e in EstadoFactura]}",
                        )

                setattr(factura, campo_modelo, valor)

        db.commit()
        db.refresh(factura)

        # Crear respuesta
        factura_response = FacturaOut(
            id_factura=factura.id_factura,
            numero_factura=factura.numero_factura,
            id_contrato=factura.id_contrato,
            fecha_emision=factura.fecha_emision,
            fecha_vencimiento=factura.fecha_vencimiento,
            subtotal=float(factura.subtotal) if factura.subtotal is not None else None,
            impuestos=(
                float(factura.impuestos) if factura.impuestos is not None else None
            ),
            descuentos=(
                float(factura.descuentos) if factura.descuentos is not None else None
            ),
            total_factura=(
                float(factura.total_factura) if factura.total_factura else 0.0
            ),
            estado_factura=(
                factura.estado_factura.value
                if hasattr(factura.estado_factura, "value")
                else factura.estado_factura
            ),
            observaciones=factura.observaciones,
            pagos=[],
        )

        return Response[FacturaOut](
            data=factura_response,
            status_code=HTTPStatus.OK,
            message="Factura actualizada exitosamente",
            error=None,
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor al actualizar factura: {str(e)}",
        )


@router.get("/facturas", response_model=Response[List[FacturaOut]])
def get_all_facturas(
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[FacturaOut]]:
    facturas = db.query(Facturas).all()
    facturas_out = []
    for factura in facturas:
        # Obtener pagos de la factura
        pagos = db.query(Pagos).filter(Pagos.id_factura == factura.id_factura).all()
        pagos_response = [
            PaymentResponseDTO(
                id_pago=pago.id_pago,
                id_factura=pago.id_factura,
                id_metodo_pago=pago.id_metodo_pago,
                id_tipo_pago=pago.id_tipo_pago,
                fecha_pago=pago.fecha_pago,
                valor=float(pago.valor) if pago.valor else 0.0,
            )
            for pago in pagos
        ]

        facturas_out.append(
            FacturaOut(
                id_factura=factura.id_factura,
                numero_factura=factura.numero_factura,
                id_contrato=factura.id_contrato,
                fecha_emision=factura.fecha_emision,
                fecha_vencimiento=factura.fecha_vencimiento,
                subtotal=(
                    float(factura.subtotal) if factura.subtotal is not None else None
                ),
                impuestos=(
                    float(factura.impuestos) if factura.impuestos is not None else None
                ),
                descuentos=(
                    float(factura.descuentos)
                    if factura.descuentos is not None
                    else None
                ),
                total_factura=(
                    float(factura.total_factura)
                    if factura.total_factura is not None
                    else None
                ),
                estado_factura=(
                    factura.estado_factura.value
                    if hasattr(factura.estado_factura, "value")
                    else factura.estado_factura
                ),
                observaciones=factura.observaciones,
                pagos=pagos_response,
            )
        )
    return Response[List[FacturaOut]](
        data=facturas_out,
        status_code=200,
        message=f"Se encontraron {len(facturas_out)} facturas",
        error=None,
    )


@router.get(
    "/contratos/{contrato_id}/facturas", response_model=Response[List[FacturaOut]]
)
def read_facturas_by_contrato(
    contrato_id: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[List[FacturaOut]]:
    """
    Obtiene todas las facturas asociadas a un contrato específico

    Args:
        contrato_id: ID del contrato
        db: Sesión de base de datos
        _: Usuario autenticado

    Returns:
        Response con lista de facturas del contrato

    Raises:
        HTTPException: Si el contrato no existe o hay errores internos
    """
    try:
        # Validar que el contrato existe
        contrato = (
            db.query(Contratos).filter(Contratos.id_contrato == contrato_id).first()
        )

        if not contrato:
            raise HTTPException(
                status_code=404, detail=f"Contrato con ID {contrato_id} no encontrado"
            )

        # Obtener facturas del contrato
        facturas = (
            db.query(Facturas)
            .filter(Facturas.id_contrato == contrato_id)
            .order_by(Facturas.fecha_creacion.desc())
            .all()
        )

        facturas_response = [
            FacturaOut(
                id_factura=factura.id_factura,
                numero_factura=factura.numero_factura,
                id_contrato=factura.id_contrato,
                fecha_emision=factura.fecha_emision,
                fecha_vencimiento=factura.fecha_vencimiento,
                subtotal=(
                    float(factura.subtotal) if factura.subtotal is not None else None
                ),
                impuestos=(
                    float(factura.impuestos) if factura.impuestos is not None else None
                ),
                descuentos=(
                    float(factura.descuentos)
                    if factura.descuentos is not None
                    else None
                ),
                total_factura=(
                    float(factura.total_factura) if factura.total_factura else 0.0
                ),
                estado_factura=(
                    factura.estado_factura.value
                    if hasattr(factura.estado_factura, "value")
                    else factura.estado_factura
                ),
                observaciones=factura.observaciones,
                pagos=[
                    PaymentResponseDTO(
                        id_pago=pago.id_pago,
                        id_factura=pago.id_factura,
                        id_metodo_pago=pago.id_metodo_pago,
                        id_tipo_pago=pago.id_tipo_pago,
                        fecha_pago=pago.fecha_pago,
                        valor=float(pago.valor) if pago.valor else 0.0,
                    )
                    for pago in db.query(Pagos)
                    .filter(Pagos.id_factura == factura.id_factura)
                    .all()
                ],
            )
            for factura in facturas
        ]

        return Response[List[FacturaOut]](
            data=facturas_response,
            status_code=HTTPStatus.OK,
            message=f"Se encontraron {len(facturas_response)} facturas para el contrato {contrato_id}",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor al obtener facturas del contrato: {str(e)}",
        )


@router.post("/facturas/{contrato_id}", response_model=Response[FacturaOut])
def create_contract_bill(
    contrato_id: int,
    factura_data: FacturaCreateWithDetails = None,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[FacturaOut]:
    """
    Crea una factura automáticamente para un contrato específico usando la lógica de negocio del CRUD

    Args:
        contrato_id: ID del contrato
        factura_data: Datos de la factura (impuestos, descuentos, observaciones)
        crud: Instancia del CRUD
        _: Usuario autenticado

    Returns:
        Response con la factura creada

    Raises:
        HTTPException: Si el contrato no existe o hay errores en la creación
    """
    try:
        # Obtener el contrato para usar fecha_fin como fecha_vencimiento
        contrato = crud._get_contract_by_id(contrato_id)

        # Calcular subtotal (suma de servicios contratados)
        subtotal = crud._calculate_contract_bill_total(contrato_id)

        # Obtener datos de facturación del payload
        impuestos = float(factura_data.impuestos) if factura_data else 0
        descuentos = float(factura_data.descuentos) if factura_data else 0
        observaciones = factura_data.observaciones if factura_data else ""

        # Calcular total: subtotal + impuestos - descuentos
        total_factura = subtotal + impuestos - descuentos
        if total_factura < 0:
            total_factura = 0

        # Crear la factura con todos los campos
        bill = Facturas(
            id_contrato=contrato_id,
            fecha_emision=datetime.now().date(),
            fecha_vencimiento=contrato.fecha_fin,  # Usar fecha_fin del contrato
            subtotal=subtotal,
            impuestos=impuestos,
            descuentos=descuentos,
            total_factura=total_factura,
            estado_factura=EstadoFactura.PENDIENTE,  # Estado inicial
            observaciones=observaciones,
        )

        crud._CareLinkCrud__carelink_session.add(bill)
        crud._CareLinkCrud__carelink_session.commit()
        crud._CareLinkCrud__carelink_session.refresh(bill)

        # Generar numero_factura basado en id_factura
        numero_factura = str(bill.id_factura).zfill(4)
        bill.numero_factura = numero_factura
        crud._CareLinkCrud__carelink_session.commit()
        crud._CareLinkCrud__carelink_session.refresh(bill)

        # Construir respuesta completa
        bill_response = FacturaOut(
            id_factura=bill.id_factura,
            numero_factura=bill.numero_factura,
            id_contrato=bill.id_contrato,
            fecha_emision=bill.fecha_emision,
            fecha_vencimiento=bill.fecha_vencimiento,
            subtotal=float(bill.subtotal) if bill.subtotal is not None else None,
            impuestos=float(bill.impuestos) if bill.impuestos is not None else None,
            descuentos=float(bill.descuentos) if bill.descuentos is not None else None,
            total_factura=float(bill.total_factura) if bill.total_factura else 0.0,
            estado_factura=(
                bill.estado_factura.value
                if hasattr(bill.estado_factura, "value")
                else bill.estado_factura
            ),
            observaciones=bill.observaciones,
        )

        return Response[FacturaOut](
            data=bill_response,
            status_code=HTTPStatus.CREATED,
            message="Factura creada automáticamente para el contrato",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al crear factura para el contrato: {str(e)}"
        )


@router.get("/facturacion/completa")
def get_facturacion_completa(db: Session = Depends(get_carelink_db)):
    # Query para facturas de contratos
    sql_contracts = text(
        """
        SELECT
            f.id_factura,
            f.numero_factura,
            f.id_contrato,
            c.tipo_contrato,
            c.fecha_inicio,
            c.fecha_fin,
            u.id_usuario,
            u.nombres,
            u.apellidos,
            u.n_documento,
            f.fecha_emision,
            f.fecha_vencimiento,
            f.total_factura,
            f.subtotal,
            f.impuestos,
            f.descuentos,
            f.estado_factura,
            f.observaciones,
            f.fecha_creacion,
            f.fecha_actualizacion,
            COUNT(p.id_pago) AS cantidad_pagos,
            COALESCE(SUM(p.valor), 0) AS total_pagado,
            'CONTRATO' as tipo_factura,
            NULL as id_visita_domiciliaria
        FROM
            Facturas f
            LEFT JOIN Contratos c ON f.id_contrato = c.id_contrato
            LEFT JOIN Usuarios u ON c.id_usuario = u.id_usuario
            LEFT JOIN Pagos p ON f.id_factura = p.id_factura
        WHERE f.id_contrato IS NOT NULL
        GROUP BY
            f.id_factura,
            f.numero_factura,
            f.id_contrato,
            c.tipo_contrato,
            c.fecha_inicio,
            c.fecha_fin,
            u.id_usuario,
            u.nombres,
            u.apellidos,
            u.n_documento,
            f.fecha_emision,
            f.fecha_vencimiento,
            f.total_factura,
            f.subtotal,
            f.impuestos,
            f.descuentos,
            f.estado_factura,
            f.observaciones,
            f.fecha_creacion,
            f.fecha_actualizacion
    """
    )

    # Query para facturas de visitas domiciliarias
    sql_home_visits = text(
        """
        SELECT
            f.id_factura,
            f.numero_factura,
            NULL as id_contrato,
            'Visita Domiciliaria' as tipo_contrato,
            NULL as fecha_inicio,
            NULL as fecha_fin,
            u.id_usuario,
            u.nombres,
            u.apellidos,
            u.n_documento,
            f.fecha_emision,
            f.fecha_vencimiento,
            f.total_factura,
            f.subtotal,
            f.impuestos,
            f.descuentos,
            f.estado_factura,
            f.observaciones,
            f.fecha_creacion,
            f.fecha_actualizacion,
            COUNT(p.id_pago) AS cantidad_pagos,
            COALESCE(SUM(p.valor), 0) AS total_pagado,
            'VISITA_DOMICILIARIA' as tipo_factura,
            f.id_visita_domiciliaria
        FROM
            Facturas f
            LEFT JOIN VisitasDomiciliarias vd ON f.id_visita_domiciliaria = vd.id_visitadomiciliaria
            LEFT JOIN Usuarios u ON vd.id_usuario = u.id_usuario
            LEFT JOIN Pagos p ON f.id_factura = p.id_factura
        WHERE f.id_visita_domiciliaria IS NOT NULL
        GROUP BY
            f.id_factura,
            f.numero_factura,
            u.id_usuario,
            u.nombres,
            u.apellidos,
            u.n_documento,
            f.fecha_emision,
            f.fecha_vencimiento,
            f.total_factura,
            f.subtotal,
            f.impuestos,
            f.descuentos,
            f.estado_factura,
            f.observaciones,
            f.fecha_creacion,
            f.fecha_actualizacion,
            f.id_visita_domiciliaria
    """
    )

    # Ejecutar ambas queries y combinar resultados
    result_contracts = db.execute(sql_contracts)
    result_home_visits = db.execute(sql_home_visits)

    # Convertir a listas de diccionarios
    contracts_rows = [dict(row) for row in result_contracts.mappings()]
    home_visits_rows = [dict(row) for row in result_home_visits.mappings()]

    # Combinar y ordenar por fecha de creación
    all_rows = contracts_rows + home_visits_rows
    all_rows.sort(key=lambda x: x.get("fecha_creacion", ""), reverse=True)

    return {"data": all_rows}


@router.get("/tarifas-servicios", response_model=Response[TarifasServicioResponseDTO])
async def get_all_service_rates(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[TarifasServicioResponseDTO]:
    """
    Obtener todas las tarifas de servicios por año con información del servicio
    """
    try:
        tarifas = crud.get_all_service_rates()

        tarifas_response = []
        for tarifa in tarifas:
            # Obtener nombre del servicio
            servicio = (
                crud._CareLinkCrud__carelink_session.query(Servicios)
                .filter(Servicios.id_servicio == tarifa.id_servicio)
                .first()
            )

            nombre_servicio = servicio.nombre if servicio else "Servicio no encontrado"

            tarifas_response.append(
                TarifaServicioResponseDTO(
                    id=tarifa.id,
                    id_servicio=tarifa.id_servicio,
                    anio=(
                        tarifa.anio.year
                        if hasattr(tarifa.anio, "year")
                        else tarifa.anio
                    ),
                    precio_por_dia=tarifa.precio_por_dia,
                    nombre_servicio=nombre_servicio,
                )
            )

        response_data = TarifasServicioResponseDTO(
            TarifasServicioPorAnio=tarifas_response
        )

        return Response[TarifasServicioResponseDTO](
            data=response_data,
            status_code=HTTPStatus.OK,
            message=f"Se obtuvieron {len(tarifas_response)} tarifas de servicios",
            error=None,
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al obtener tarifas de servicios: {str(e)}"
        )


@router.patch("/tarifas-servicios", response_model=Response[TarifasServicioResponseDTO])
async def update_service_rates(
    tarifas_data: TarifasServicioUpdateRequestDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
) -> Response[TarifasServicioResponseDTO]:
    """
    Actualizar múltiples tarifas de servicios por año
    """
    try:
        # Convertir DTOs a diccionarios para el CRUD
        tarifas_dict = [
            {
                "id": tarifa.id,
                "id_servicio": tarifa.id_servicio,
                "anio": tarifa.anio,
                "precio_por_dia": float(tarifa.precio_por_dia),
            }
            for tarifa in tarifas_data.TarifasServicioPorAnio
        ]

        # Actualizar tarifas
        updated_tarifas = crud.update_service_rates(tarifas_dict)

        # Construir respuesta
        tarifas_response = []
        for tarifa in updated_tarifas:
            # Obtener nombre del servicio
            servicio = (
                crud._CareLinkCrud__carelink_session.query(Servicios)
                .filter(Servicios.id_servicio == tarifa.id_servicio)
                .first()
            )

            nombre_servicio = servicio.nombre if servicio else "Servicio no encontrado"

            tarifas_response.append(
                TarifaServicioResponseDTO(
                    id=tarifa.id,
                    id_servicio=tarifa.id_servicio,
                    anio=(
                        tarifa.anio.year
                        if hasattr(tarifa.anio, "year")
                        else tarifa.anio
                    ),
                    precio_por_dia=tarifa.precio_por_dia,
                    nombre_servicio=nombre_servicio,
                )
            )

        response_data = TarifasServicioResponseDTO(
            TarifasServicioPorAnio=tarifas_response
        )

        return Response[TarifasServicioResponseDTO](
            data=response_data,
            status_code=HTTPStatus.OK,
            message=f"Se actualizaron {len(tarifas_response)} tarifas de servicios exitosamente",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al actualizar tarifas de servicios: {str(e)}",
        )


@router.get("/facturas/estadisticas")
def get_facturas_estadisticas(
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    """
    Obtiene estadísticas calculadas de facturación
    """
    try:
        # Obtener todas las facturas con sus pagos
        facturas = db.query(Facturas).all()

        total_facturas = len(facturas)
        total_valor = 0
        valor_pendiente = 0
        valor_pagado = 0
        pagadas = 0
        pendientes = 0
        vencidas = 0
        canceladas = 0
        anuladas = 0

        for factura in facturas:
            total_factura = float(factura.total_factura) if factura.total_factura else 0
            total_valor += total_factura

            # Calcular total pagado
            pagos = db.query(Pagos).filter(Pagos.id_factura == factura.id_factura).all()
            total_pagado = sum(float(pago.valor) for pago in pagos if pago.valor)
            valor_pagado += total_pagado
            valor_pendiente += max(0, total_factura - total_pagado)

            # Contar por estado
            estado = (
                factura.estado_factura.value
                if hasattr(factura.estado_factura, "value")
                else factura.estado_factura
            )
            if estado == "PAGADA":
                pagadas += 1
            elif estado == "PENDIENTE":
                pendientes += 1
            elif estado == "VENCIDA":
                vencidas += 1
            elif estado == "CANCELADA":
                canceladas += 1
            elif estado == "ANULADA":
                anuladas += 1

        # Calcular porcentajes
        porcentaje_pagadas = (
            (pagadas / total_facturas * 100) if total_facturas > 0 else 0
        )
        porcentaje_valor_pagado = (
            (valor_pagado / total_valor * 100) if total_valor > 0 else 0
        )
        promedio_por_factura = total_valor / total_facturas if total_facturas > 0 else 0

        return {
            "data": {
                "total_facturas": total_facturas,
                "facturas_pagadas": pagadas,
                "facturas_pendientes": pendientes,
                "facturas_vencidas": vencidas,
                "facturas_canceladas": canceladas,
                "facturas_anuladas": anuladas,
                "valor_total": total_valor,
                "valor_pagado": valor_pagado,
                "valor_pendiente": valor_pendiente,
                "promedio_por_factura": promedio_por_factura,
                "porcentaje_pagadas": round(porcentaje_pagadas, 1),
                "porcentaje_valor_pagado": round(porcentaje_valor_pagado, 1),
            }
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al calcular estadísticas de facturación: {str(e)}",
        )


@router.get(
    "/facturas/{id_factura}/pagos/total", response_model=BillPaymentsTotalResponseDTO
)
def get_bill_payments_total_endpoint(
    id_factura: int,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(
        require_roles(Role.ADMIN.value, Role.PROFESSIONAL.value)
    ),
):
    """
    Retorna el total de pagos asociados a una factura
    """
    from app.crud.carelink_crud import get_bill_payments_total

    total = get_bill_payments_total(db, id_factura)
    return BillPaymentsTotalResponseDTO(id_factura=id_factura, total_pagado=total)


@router.get("/facturas/{id_factura}/pdf")
async def generate_factura_pdf(
    id_factura: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
):
    """
    Genera un PDF de la factura con toda la información relacionada
    """
    try:
        # Obtener todos los datos necesarios
        factura_data = crud.get_complete_factura_data_for_pdf(id_factura)

        if not factura_data:
            raise HTTPException(
                status_code=404, detail=f"Factura con ID {id_factura} no encontrada"
            )

        # Generar el PDF
        pdf_bytes = crud.generate_factura_pdf(factura_data)

        # Devolver el PDF como archivo descargable
        from fastapi.responses import Response

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=factura_{id_factura}.pdf"
            },
        )

    except Exception as e:
        print(f"Error generando PDF de factura {id_factura}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")


@router.delete("/photo/{user_id}")
async def delete_user_photo(user_id: int, crud: CareLinkCrud = Depends(get_crud)):
    crud.delete_user_photo(user_id)
    raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")


@router.delete("/contratos/{id_contrato}", status_code=204)
def eliminar_contrato(id_contrato: int, db: Session = Depends(get_carelink_db)):
    """
    Elimina un contrato y desasocia las facturas (pone id_contrato en NULL).
    - Elimina cronogramas asociados con estado 'PENDIENTE'.
    - Desasocia (pone id_contrato en NULL) los demás cronogramas.
    """
    try:
        contrato = (
            db.query(Contratos).filter(Contratos.id_contrato == id_contrato).first()
        )
        if not contrato:
            raise HTTPException(status_code=404, detail="Contrato no encontrado")
        # Desasociar facturas
        db.query(Facturas).filter(Facturas.id_contrato == id_contrato).update(
            {"id_contrato": None}
        )
        # Eliminar cronogramas PENDIENTE
        db.query(CronogramaAsistenciaPacientes).filter(
            CronogramaAsistenciaPacientes.id_contrato == id_contrato,
            CronogramaAsistenciaPacientes.estado_asistencia == "PENDIENTE",
        ).delete()
        # Desasociar los demás cronogramas
        db.query(CronogramaAsistenciaPacientes).filter(
            CronogramaAsistenciaPacientes.id_contrato == id_contrato,
            CronogramaAsistenciaPacientes.estado_asistencia != "PENDIENTE",
        ).update({"id_contrato": None})
        # Eliminar el contrato
        db.delete(contrato)
        db.commit()
        return {
            "ok": True,
            "message": "Contrato eliminado, facturas desasociadas y cronogramas gestionados",
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Error al eliminar contrato: {str(e)}"
        )


# Endpoints para visitas domiciliarias
@router.get(
    "/users/{user_id}/home-visits",
    response_model=Response[List[VisitaDomiciliariaConProfesionalResponseDTO]],
)
async def get_user_home_visits(
    user_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[VisitaDomiciliariaConProfesionalResponseDTO]]:
    """Obtener todas las visitas domiciliarias de un usuario"""
    try:
        visitas = crud.get_home_visits_with_professionals(user_id)
        # Convertir los datos del CRUD al DTO para asegurar la serialización correcta
        visitas_dto = [
            VisitaDomiciliariaConProfesionalResponseDTO(**visita) for visita in visitas
        ]
        return Response(
            data=visitas_dto,
            message="Visitas domiciliarias obtenidas exitosamente",
            status_code=200,
            error=None,
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al obtener visitas domiciliarias: {str(e)}"
        )


@router.get(
    "/home-visits",
    response_model=Response[List[VisitaDomiciliariaConProfesionalResponseDTO]],
)
async def get_all_home_visits(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[VisitaDomiciliariaConProfesionalResponseDTO]]:
    """Obtener todas las visitas domiciliarias (para permitir filtrado completo)"""
    try:
        visitas = crud.get_all_home_visits_with_professionals()
        # Convertir los datos del CRUD al DTO para asegurar la serialización correcta
        visitas_dto = [
            VisitaDomiciliariaConProfesionalResponseDTO(**visita) for visita in visitas
        ]
        return Response(
            data=visitas_dto,
            message="Visitas domiciliarias obtenidas exitosamente",
            status_code=200,
            error=None,
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al obtener visitas domiciliarias: {str(e)}"
        )


@router.get(
    "/home-visits/all",
    response_model=Response[List[VisitaDomiciliariaConProfesionalResponseDTO]],
)
async def get_all_home_visits_history(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[VisitaDomiciliariaConProfesionalResponseDTO]]:
    """Obtener todas las visitas domiciliarias (historial completo)"""
    try:
        visitas = crud.get_all_home_visits_with_professionals()
        # Convertir los datos del CRUD al DTO para asegurar la serialización correcta
        visitas_dto = [
            VisitaDomiciliariaConProfesionalResponseDTO(**visita) for visita in visitas
        ]
        return Response(
            data=visitas_dto,
            message="Historial de visitas domiciliarias obtenido exitosamente",
            status_code=200,
            error=None,
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener historial de visitas domiciliarias: {str(e)}",
        )


@router.get(
    "/home-visits/{visita_id}", response_model=Response[VisitaDomiciliariaResponseDTO]
)
async def get_home_visit_by_id(
    visita_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[VisitaDomiciliariaResponseDTO]:
    """Obtener una visita domiciliaria por ID"""
    try:
        visita = crud.get_home_visit_by_id(visita_id)
        return Response(
            data=VisitaDomiciliariaResponseDTO.from_orm(visita),
            message="Visita domiciliaria obtenida exitosamente",
            status_code=200,
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al obtener visita domiciliaria: {str(e)}"
        )


@router.post(
    "/users/{user_id}/home-visits",
    response_model=Response[VisitaDomiciliariaResponseDTO],
)
async def create_home_visit(
    user_id: int,
    visita_data: VisitaDomiciliariaCreateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[VisitaDomiciliariaResponseDTO]:
    """Crear una nueva visita domiciliaria"""
    try:
        # Verificar que el usuario existe
        crud.list_user_by_user_id(user_id)

        # Extraer el id_profesional_asignado si existe
        id_profesional_asignado = visita_data.id_profesional_asignado

        # Crear la visita
        visita_dict = visita_data.dict()
        visita_dict["id_usuario"] = user_id

        # Remover id_profesional_asignado del dict para no incluirlo en la tabla VisitasDomiciliarias
        visita_dict.pop("id_profesional_asignado", None)

        visita = crud.create_home_visit_manual(visita_dict, id_profesional_asignado)

        return Response(
            data=VisitaDomiciliariaResponseDTO.from_orm(visita),
            message="Visita domiciliaria creada exitosamente",
            status_code=201,
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al crear visita domiciliaria: {str(e)}"
        )


@router.patch(
    "/home-visits/{visita_id}", response_model=Response[VisitaDomiciliariaResponseDTO]
)
async def update_home_visit(
    visita_id: int,
    visita_data: VisitaDomiciliariaUpdateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[VisitaDomiciliariaResponseDTO]:
    """Actualizar una visita domiciliaria"""
    try:
        # Filtrar solo los campos que no son None
        update_data = {k: v for k, v in visita_data.dict().items() if v is not None}

        visita = crud.update_home_visit(visita_id, update_data)

        return Response(
            data=VisitaDomiciliariaResponseDTO.from_orm(visita),
            message="Visita domiciliaria actualizada exitosamente",
            status_code=200,
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al actualizar visita domiciliaria: {str(e)}"
        )


@router.delete("/home-visits/{visita_id}", response_model=Response[dict])
async def delete_home_visit(
    visita_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[dict]:
    """Eliminar una visita domiciliaria"""
    try:
        crud.delete_home_visit(visita_id)

        return Response(
            data={"deleted": True},
            message="Visita domiciliaria eliminada exitosamente",
            success=True,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al eliminar visita domiciliaria: {str(e)}"
        )


@router.put(
    "/users/{user_id}/home-visits/{visita_id}",
    response_model=Response[VisitaDomiciliariaResponseDTO],
)
async def update_user_home_visit(
    user_id: int,
    visita_id: int,
    visita_data: VisitaDomiciliariaUpdateDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[VisitaDomiciliariaResponseDTO]:
    """Actualizar una visita domiciliaria específica de un usuario"""
    try:
        # Verificar que la visita pertenece al usuario
        visita = crud.get_home_visit_by_id(visita_id)
        if visita.id_usuario != user_id:
            raise HTTPException(
                status_code=403, detail="La visita no pertenece al usuario especificado"
            )

        # Filtrar solo los campos que no son None
        update_data = {k: v for k, v in visita_data.dict().items() if v is not None}

        visita_actualizada = crud.update_home_visit(visita_id, update_data)
        return Response(
            data=VisitaDomiciliariaResponseDTO.from_orm(visita_actualizada),
            message="Visita domiciliaria actualizada exitosamente",
            status_code=200,
            error=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al actualizar visita domiciliaria: {str(e)}"
        )


@router.get(
    "/asistencia/diaria", response_model=Response[List[AsistenciaDiariaResponseDTO]]
)
def get_asistencia_diaria(
    fecha: Optional[str] = None,
    db: Session = Depends(get_carelink_db),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[AsistenciaDiariaResponseDTO]]:
    """
    Obtiene la asistencia del día actual (o fecha especificada) para el dashboard
    """
    try:
        # Si no se especifica fecha, usar la fecha actual
        if fecha:
            fecha_consulta = datetime.strptime(fecha, "%Y-%m-%d").date()
        else:
            fecha_consulta = date.today()

        # Consultar cronogramas para la fecha especificada
        cronogramas = (
            db.query(CronogramaAsistencia)
            .filter(CronogramaAsistencia.fecha == fecha_consulta)
            .all()
        )

        result = []
        for cronograma in cronogramas:
            # Obtener pacientes agendados para este cronograma con información completa
            pacientes_agendados = (
                db.query(CronogramaAsistenciaPacientes, User, Contratos)
                .join(User, CronogramaAsistenciaPacientes.id_usuario == User.id_usuario)
                .join(
                    Contratos,
                    CronogramaAsistenciaPacientes.id_contrato == Contratos.id_contrato,
                )
                .filter(
                    CronogramaAsistenciaPacientes.id_cronograma
                    == cronograma.id_cronograma
                )
                .all()
            )

            for paciente_agendado, usuario, contrato in pacientes_agendados:
                # Determinar el tipo de servicio basado en el contrato
                tipo_servicio = contrato.tipo_contrato if contrato else "Sin servicio"

                # Mapear estado de asistencia a texto legible
                estado_texto = {
                    "PENDIENTE": "Pendiente",
                    "ASISTIO": "Asistió",
                    "NO_ASISTIO": "No asistió",
                    "CANCELADO": "Cancelado",
                    "REAGENDADO": "Reagendado",
                }.get(
                    paciente_agendado.estado_asistencia,
                    paciente_agendado.estado_asistencia,
                )

                # Color del estado
                color_estado = {
                    "PENDIENTE": "gray",
                    "ASISTIO": "green",
                    "NO_ASISTIO": "red",
                    "CANCELADO": "orange",
                    "REAGENDADO": "blue",
                }.get(paciente_agendado.estado_asistencia, "default")

                result.append(
                    AsistenciaDiariaResponseDTO(
                        id_cronograma_paciente=paciente_agendado.id_cronograma_paciente,
                        id_usuario=usuario.id_usuario,
                        nombres=usuario.nombres,
                        apellidos=usuario.apellidos,
                        tipo_servicio=tipo_servicio,
                        estado_asistencia=paciente_agendado.estado_asistencia,
                        estado_texto=estado_texto,
                        color_estado=color_estado,
                        requiere_transporte=paciente_agendado.requiere_transporte,
                        observaciones=paciente_agendado.observaciones,
                        fecha_creacion=paciente_agendado.fecha_creacion,
                        fecha_actualizacion=paciente_agendado.fecha_actualizacion,
                    )
                )

        return Response[List[AsistenciaDiariaResponseDTO]](
            data=result,
            status_code=HTTPStatus.OK,
            message=f"Asistencia del día {fecha_consulta} consultada exitosamente",
            error=None,
        )

    except ValueError as e:
        raise HTTPException(
            status_code=400, detail=f"Formato de fecha inválido: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.get(
    "/pagos/factura/{factura_id}", response_model=Response[List[PaymentResponseDTO]]
)
async def get_pagos_by_factura(
    factura_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[PaymentResponseDTO]]:
    """Obtiene todos los pagos de una factura específica"""
    try:
        # Verificar que la factura existe
        bill = crud.get_bill_by_id(factura_id)
        if not bill:
            raise HTTPException(
                status_code=404, detail=f"La factura con ID {factura_id} no existe"
            )

        # Obtener pagos de la factura
        payments = crud.get_payments_by_factura(factura_id)
        payments_response = [
            PaymentResponseDTO.from_orm(payment) for payment in payments
        ]

        return Response[List[PaymentResponseDTO]](
            data=payments_response,
            status_code=HTTPStatus.OK,
            message=f"Pagos de la factura {factura_id} obtenidos exitosamente",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.post("/facturas/{factura_id}/pagos/", response_model=Response[dict])
async def add_pagos_to_factura(
    factura_id: int,
    pagos: List[PagoCreate],
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[dict]:
    """Agrega múltiples pagos a una factura específica"""
    try:
        # Verificar que la factura existe
        bill = crud.get_bill_by_id(factura_id)
        if not bill:
            raise HTTPException(
                status_code=404, detail=f"La factura con ID {factura_id} no existe"
            )

        # Validar métodos y tipos de pago
        payment_methods = crud._get_payment_methods()
        payment_types = crud._get_payment_types()

        for pago in pagos:
            # Validar método de pago
            if not any(
                pm.id_metodo_pago == pago.id_metodo_pago for pm in payment_methods
            ):
                raise HTTPException(
                    status_code=400,
                    detail=f"El método de pago con ID {pago.id_metodo_pago} no existe",
                )

            # Validar tipo de pago
            if not any(pt.id_tipo_pago == pago.id_tipo_pago for pt in payment_types):
                raise HTTPException(
                    status_code=400,
                    detail=f"El tipo de pago con ID {pago.id_tipo_pago} no existe",
                )

        created_payments = []
        for pago_data in pagos:
            payment = Pagos(
                id_factura=factura_id,
                id_metodo_pago=pago_data.id_metodo_pago,
                id_tipo_pago=pago_data.id_tipo_pago,
                fecha_pago=pago_data.fecha_pago,
                valor=pago_data.valor,
            )
            created_payment = crud.create_payment(payment)
            created_payments.append(created_payment)

        # Actualizar el estado de la factura
        crud.update_factura_status(factura_id)

        return Response[dict](
            data={
                "message": f"Se agregaron {len(created_payments)} pagos a la factura {factura_id}",
                "pagos_creados": len(created_payments),
                "factura_id": factura_id,
            },
            status_code=HTTPStatus.CREATED,
            message="Pagos agregados exitosamente",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error interno del servidor: {str(e)}"
        )


@router.get("/user-flow", response_model=Response[UserFlowResponseDTO])
async def get_user_flow(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[UserFlowResponseDTO]:
    """
    Obtiene los datos del flujo de usuarios para el dashboard.
    Incluye estadísticas y lista de usuarios con visitas domiciliarias = false.
    """
    try:
        result = crud.get_user_flow_data()
        return Response[UserFlowResponseDTO](
            data=result,
            status_code=HTTPStatus.OK,
            message="Datos del flujo de usuarios obtenidos exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[UserFlowResponseDTO](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al obtener datos del flujo de usuarios: {str(e)}",
            error=None,
        )


@router.get("/quarterly-visits", response_model=Response[QuarterlyVisitsResponseDTO])
async def get_quarterly_visits(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[QuarterlyVisitsResponseDTO]:
    """
    Obtiene los datos de visitas del trimestre para el dashboard.
    Incluye estadísticas trimestrales y datos mensuales para el gráfico.
    """
    try:
        result = crud.get_quarterly_visits_data()
        return Response[QuarterlyVisitsResponseDTO](
            data=result,
            status_code=HTTPStatus.OK,
            message="Datos de visitas trimestrales obtenidos exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[QuarterlyVisitsResponseDTO](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al obtener datos de visitas trimestrales: {str(e)}",
            error=None,
        )


@router.get("/monthly-payments", response_model=Response[MonthlyPaymentsResponseDTO])
async def get_monthly_payments(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[MonthlyPaymentsResponseDTO]:
    """
    Obtiene los datos de pagos mensuales para el dashboard.
    Incluye estadísticas de pagos y metas basadas en el mes anterior.
    """
    try:
        result = crud.get_monthly_payments_data()
        return Response[MonthlyPaymentsResponseDTO](
            data=result,
            status_code=HTTPStatus.OK,
            message="Datos de pagos mensuales obtenidos exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[MonthlyPaymentsResponseDTO](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al obtener datos de pagos mensuales: {str(e)}",
            error=None,
        )


@router.get(
    "/operational-efficiency", response_model=Response[OperationalEfficiencyResponseDTO]
)
async def get_operational_efficiency(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[OperationalEfficiencyResponseDTO]:
    """
    Obtiene los datos de eficiencia operativa para el dashboard.
    Combina múltiples métricas: asistencia, visitas domiciliarias, contratos y facturación.
    """
    try:
        result = crud.get_operational_efficiency_data()
        return Response[OperationalEfficiencyResponseDTO](
            data=result,
            status_code=HTTPStatus.OK,
            message="Datos de eficiencia operativa obtenidos exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[OperationalEfficiencyResponseDTO](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al obtener datos de eficiencia operativa: {str(e)}",
            error=None,
        )


# ============================================================================
# ENDPOINTS PARA GESTIÓN DE USUARIOS EN ACTIVIDADES
# ============================================================================


@router.get(
    "/activities/{activity_id}/users", response_model=Response[ActivityWithUsersDTO]
)
async def get_activity_with_users(
    activity_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[ActivityWithUsersDTO]:
    """Obtener una actividad con sus usuarios asignados"""
    try:
        result = crud.get_activity_with_users(activity_id)
        return Response[ActivityWithUsersDTO](
            data=result,
            status_code=HTTPStatus.OK,
            message="Actividad con usuarios obtenida exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[ActivityWithUsersDTO](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al obtener actividad con usuarios: {str(e)}",
            error=None,
        )


@router.get(
    "/activities/users/available/{activity_date}",
    response_model=Response[List[UserForActivityDTO]],
)
async def get_users_for_activity_date(
    activity_date: str,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[List[UserForActivityDTO]]:
    """Obtener usuarios disponibles para una fecha específica basado en el cronograma"""
    try:
        from datetime import datetime

        date_obj = datetime.strptime(activity_date, "%Y-%m-%d").date()
        result = crud.get_users_for_activity_date(date_obj)

        return Response[List[UserForActivityDTO]](
            data=result,
            status_code=HTTPStatus.OK,
            message="Usuarios disponibles obtenidos exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[List[UserForActivityDTO]](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al obtener usuarios disponibles: {str(e)}",
            error=None,
        )


@router.get("/users/{user_id}/download-contract/{contract_type}")
async def download_contract(
    user_id: int,
    contract_type: str,
    quantity: Optional[int] = Query(
        None, description="Cantidad de días para el contrato"
    ),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
):
    """Descargar contrato de Word con información renderizada"""
    try:
        # Obtener información del usuario
        user = crud._get_user_by_id(user_id)
        if not user:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        # Obtener información del registro médico del usuario para EPS
        medical_record = crud._get_user_medical_record_by_user_id(user_id)
        eps_info = medical_record.eps if medical_record else "No especificado"

        # Obtener información del acudiente
        guardian_info = crud._get_user_guardian_info(user_id)

        # Calcular edad
        from datetime import datetime

        birth_date = datetime.strptime(str(user.fecha_nacimiento), "%Y-%m-%d")
        age = datetime.now().year - birth_date.year
        if datetime.now().month < birth_date.month or (
            datetime.now().month == birth_date.month
            and datetime.now().day < birth_date.day
        ):
            age -= 1

        # Preparar datos para el template
        # Usar la cantidad proporcionada o un valor por defecto
        dias_por_tiquetera = str(quantity) if quantity is not None else "20"

        # Obtener precios dinámicos según el tipo de contrato
        current_year = datetime.now().year
        valor_dia = 0.0
        valor_total = 0.0

        if contract_type == "transporte":
            # Obtener precio del servicio de transporte
            valor_dia = crud._get_service_price_by_name("transporte", current_year)
            if quantity and valor_dia > 0:
                valor_total = valor_dia * quantity
            else:
                valor_dia = 50000  # Valor por defecto
                valor_total = valor_dia * (quantity or 20)
        else:
            # Para centro de día, usar valores por defecto
            valor_dia = 50000
            valor_total = valor_dia * (quantity or 20)

        context = {
            "fecha_de_impresion": datetime.now().strftime("%d/%m/%Y"),
            "ID_del_paciente_en_base_datos": user.id_usuario,
            "Nombre_Paciente": f"{user.nombres} {user.apellidos}",
            "Numero_identificación": user.n_documento,
            "Fecha_Nacimiento": user.fecha_nacimiento.strftime("%d/%m/%Y"),
            "Convenio_Con_Empresas": eps_info,
            "Sexo_Paciente": user.genero,
            "Edad_Años_Meses": f"{age} años",
            "Numero_de_días_por_tiquetera": dias_por_tiquetera,
            "Nombre_del_paciente": f"{user.nombres} {user.apellidos}",
            "Dirección_del_paciente": user.direccion or "No especificada",
            "Barrio_del_paciente": "No especificado",
            "Telefono_del_paciente": user.telefono or "No especificado",
            "Email_del_paciente": user.email or "No especificado",
            "nombre_del_acudiente": guardian_info["nombre_completo"],
            "telefono_del_acudiente": guardian_info["telefono"],
            "documento_del_acudiente": guardian_info["documento"],
            "valor_dia": f"${valor_dia:,.0f}",
            "valor_total": f"${valor_total:,.0f}",
            "fecha_de_firma": datetime.now().strftime("%d/%m/%Y"),
        }

        # Determinar qué template usar
        if contract_type == "centro-dia":
            template_path = "app/static/templates/CONTRATO CENTRO DE DIA-1752465077348-949936205 (1).docx"
        elif contract_type == "transporte":
            template_path = "app/static/templates/CONTRATO DE TRANSPORTE-1752467433320-608946576 (2).docx"
        else:
            raise HTTPException(status_code=400, detail="Tipo de contrato no válido")

        # Verificar que el archivo existe
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404, detail="Template de contrato no encontrado"
            )

        # Generar el documento
        doc = DocxTemplate(template_path)
        doc.render(context)

        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp_file:
            doc.save(tmp_file.name)
            tmp_file_path = tmp_file.name

        # Generar nombre del archivo
        filename = f"contrato_{contract_type}_{user.nombres}_{user.apellidos}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"

        return FileResponse(
            path=tmp_file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al generar contrato: {str(e)}"
        )


@router.post("/activities/{activity_id}/users", response_model=Response[dict])
async def assign_users_to_activity(
    activity_id: int,
    assign_data: AssignUsersToActivityDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[dict]:
    """Asignar usuarios a una actividad"""
    try:
        result = crud.assign_users_to_activity(
            activity_id=activity_id,
            user_ids=assign_data.usuarios_ids,
            estado_participacion=assign_data.estado_participacion,
            observaciones=assign_data.observaciones,
        )
        return Response[dict](
            data={"message": f"Se asignaron {len(result)} usuarios a la actividad"},
            status_code=HTTPStatus.OK,
            message="Usuarios asignados exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[dict](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al asignar usuarios: {str(e)}",
            error=None,
        )


@router.delete("/activities/{activity_id}/users", response_model=Response[dict])
async def remove_users_from_activity(
    activity_id: int,
    user_ids: List[int],
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[dict]:
    """Remover usuarios de una actividad"""
    try:
        result = crud.remove_users_from_activity(activity_id, user_ids)
        return Response[dict](
            data={"message": f"Se removieron {len(user_ids)} usuarios de la actividad"},
            status_code=HTTPStatus.OK,
            message="Usuarios removidos exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[dict](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al remover usuarios: {str(e)}",
            error=None,
        )


@router.patch(
    "/activities/users/{activity_user_id}/status", response_model=Response[dict]
)
async def update_user_activity_status(
    activity_user_id: int,
    status_data: UpdateUserActivityStatusDTO,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[dict]:
    """Actualizar el estado de participación de un usuario en una actividad"""
    try:
        result = crud.update_user_activity_status(
            activity_user_id=activity_user_id,
            estado_participacion=status_data.estado_participacion,
            observaciones=status_data.observaciones,
        )
        return Response[dict](
            data={"message": "Estado de participación actualizado"},
            status_code=HTTPStatus.OK,
            message="Estado actualizado exitosamente",
            error=None,
        )
    except Exception as e:
        return Response[dict](
            data=None,
            status_code=HTTPStatus.INTERNAL_SERVER_ERROR,
            message=f"Error al actualizar estado: {str(e)}",
            error=None,
        )


@router.post("/users/{user_id}/home-visits/bills", response_model=Response[FacturaOut])
async def create_home_visit_bill(
    user_id: int,
    bill_data: dict,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
) -> Response[FacturaOut]:
    """Crear una factura para una visita domiciliaria"""
    try:
        from datetime import datetime
        from app.models.contracts import Facturas, EstadoFactura

        # Verificar que la visita domiciliaria existe
        visita_id = bill_data.get("id_visita_domiciliaria")
        if not visita_id:
            raise HTTPException(
                status_code=400, detail="id_visita_domiciliaria es requerido"
            )

        visita = crud.get_home_visit_by_id(visita_id)
        if visita.id_usuario != user_id:
            raise HTTPException(
                status_code=403, detail="La visita no pertenece al usuario especificado"
            )

        # Generar número de factura secuencial
        current_year = datetime.now().year
        next_invoice_number = crud._get_next_home_visit_invoice_number(current_year)

        # Crear la factura
        factura = Facturas(
            id_contrato=None,  # Las visitas domiciliarias no tienen contrato
            id_visita_domiciliaria=visita_id,
            fecha_emision=bill_data.get("fecha_emision", datetime.now().date()),
            fecha_vencimiento=bill_data.get("fecha_vencimiento"),
            subtotal=bill_data.get("subtotal", 0),
            impuestos=bill_data.get("impuestos", 0),
            descuentos=bill_data.get("descuentos", 0),
            total_factura=bill_data.get("total_factura", 0),
            estado_factura=EstadoFactura.PENDIENTE,
            observaciones=bill_data.get("observaciones", ""),
            numero_factura=next_invoice_number,
        )

        crud._CareLinkCrud__carelink_session.add(factura)
        crud._CareLinkCrud__carelink_session.commit()
        crud._CareLinkCrud__carelink_session.refresh(factura)

        # Crear respuesta
        factura_response = FacturaOut(
            id_factura=factura.id_factura,
            numero_factura=factura.numero_factura,
            id_contrato=factura.id_contrato,
            fecha_emision=factura.fecha_emision,
            fecha_vencimiento=factura.fecha_vencimiento,
            subtotal=float(factura.subtotal) if factura.subtotal is not None else None,
            impuestos=(
                float(factura.impuestos) if factura.impuestos is not None else None
            ),
            descuentos=(
                float(factura.descuentos) if factura.descuentos is not None else None
            ),
            total_factura=(
                float(factura.total_factura) if factura.total_factura else 0.0
            ),
            estado_factura=(
                factura.estado_factura.value
                if hasattr(factura.estado_factura, "value")
                else factura.estado_factura
            ),
            observaciones=factura.observaciones,
        )

        return Response[FacturaOut](
            data=factura_response,
            status_code=HTTPStatus.CREATED,
            message="Factura de visita domiciliaria creada exitosamente",
            error=None,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al crear factura de visita domiciliaria: {str(e)}",
        )


@router.get("/users/template/excel")
async def export_user_template(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
):
    """
    Genera y descarga una plantilla Excel para importación masiva de usuarios.
    Solo incluye usuarios que NO están relacionados con visitas domiciliarias.
    """
    try:
        # Crear un nuevo workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Plantilla Usuarios"

        # Definir columnas
        columns = [
            "Tipo de usuario",
            "N° Documento",
            "Nombres",
            "Apellidos",
            "Género",
            "Fecha de nacimiento",
            "Estado civil",
            "Ocupación",
        ]

        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Escribir encabezados
        for col_num, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Ajustar ancho de columnas
        column_widths = [15, 15, 20, 20, 12, 15, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + col_num)].width = width

        # Agregar datos de ejemplo
        example_data = [
            [
                "Nuevo",
                "1234567890",
                "Juan",
                "Pérez",
                "Masculino",
                "1990-01-15",
                "Soltero",
                "Ingeniero",
            ],
            [
                "Recurrente",
                "0987654321",
                "María",
                "García",
                "Femenino",
                "1985-05-20",
                "Casado",
                "Médico",
            ],
            [
                "Nuevo",
                "1122334455",
                "Carlos",
                "López",
                "Masculino",
                "1995-12-10",
                "Soltero",
                "Abogado",
            ],
        ]

        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.border = border

        # Agregar instrucciones
        worksheet.cell(row=6, column=1, value="INSTRUCCIONES:")
        worksheet.cell(
            row=7, column=1, value="1. Complete los datos en las filas correspondientes"
        )
        worksheet.cell(
            row=8, column=1, value="2. Tipo de usuario: 'Nuevo' o 'Recurrente'"
        )
        worksheet.cell(
            row=9, column=1, value="3. Género: 'Masculino', 'Femenino' o 'Neutro'"
        )
        worksheet.cell(
            row=10,
            column=1,
            value="4. Estado civil: 'Soltero', 'Casado', 'Divorciado', 'Viudo', 'Unión Libre'",
        )
        worksheet.cell(
            row=11, column=1, value="5. Fecha de nacimiento: formato YYYY-MM-DD"
        )
        worksheet.cell(
            row=12, column=1, value="6. Los campos marcados con * son obligatorios"
        )

        # Guardar en un archivo temporal
        import tempfile
        import os

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            workbook.save(tmp_file.name)
            tmp_file_path = tmp_file.name

        return FileResponse(
            tmp_file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="plantilla_usuarios_fundacion.xlsx",
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar la plantilla Excel: {str(e)}",
        )


@router.post("/users/import/excel")
async def import_users_from_excel(
    file: UploadFile = File(...),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
):
    """
    Importa usuarios masivamente desde un archivo Excel.
    Solo crea usuarios para asistencia a la fundación (NO visitas domiciliarias).
    """
    from datetime import datetime

    try:
        # Validar tipo de archivo
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo debe ser un archivo Excel (.xlsx)",
            )

        # Leer el archivo Excel con openpyxl
        workbook = load_workbook(file.file)
        worksheet = workbook.active

        # Obtener encabezados de la primera fila
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value)

        # Validar columnas requeridas
        required_columns = [
            "Tipo de usuario",
            "N° Documento",
            "Nombres",
            "Apellidos",
            "Género",
            "Fecha de nacimiento",
            "Estado civil",
            "Ocupación",
        ]

        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Columnas faltantes en el archivo: {', '.join(missing_columns)}",
            )

        # Procesar cada fila
        results = {
            "success": [],
            "errors": [],
            "total_processed": 0,
            "total_success": 0,
            "total_errors": 0,
        }

        # Procesar filas desde la segunda (saltando encabezados)
        for row_num in range(2, worksheet.max_row + 1):
            results["total_processed"] += 1

            try:
                # Obtener valores de la fila
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value

                # Validar datos requeridos
                if (
                    not row_data.get("Nombres")
                    or not row_data.get("Apellidos")
                    or not row_data.get("N° Documento")
                ):
                    results["errors"].append(
                        {
                            "row": row_num,
                            "error": "Los campos Nombres, Apellidos y N° Documento son obligatorios",
                        }
                    )
                    results["total_errors"] += 1
                    continue

                # Validar tipo de usuario
                tipo_usuario_raw = (
                    str(row_data.get("Tipo de usuario", "")).strip() or "Nuevo"
                )
                tipo_usuario = tipo_usuario_raw.title()  # Convierte "NUEVO" a "Nuevo"
                if tipo_usuario not in ["Nuevo", "Recurrente"]:
                    results["errors"].append(
                        {
                            "row": row_num,
                            "error": f"Tipo de usuario '{tipo_usuario_raw}' no válido. Debe ser 'Nuevo' o 'Recurrente'",
                        }
                    )
                    results["total_errors"] += 1
                    continue

                # Validar género
                genero_raw = str(row_data.get("Género", "")).strip() or "Masculino"
                genero = genero_raw.title()  # Convierte "MASCULINO" a "Masculino"
                if genero not in ["Masculino", "Femenino", "Neutro"]:
                    results["errors"].append(
                        {
                            "row": row_num,
                            "error": f"Género '{genero_raw}' no válido. Debe ser 'Masculino', 'Femenino' o 'Neutro'",
                        }
                    )
                    results["total_errors"] += 1
                    continue

                # Validar estado civil
                estado_civil_raw = (
                    str(row_data.get("Estado civil", "")).strip() or "Soltero"
                )
                estado_civil = (
                    estado_civil_raw.title()
                )  # Convierte "SOLTERO" a "Soltero"
                if estado_civil not in [
                    "Soltero",
                    "Casado",
                    "Divorciado",
                    "Viudo",
                    "Unión Libre",
                ]:
                    results["errors"].append(
                        {
                            "row": row_num,
                            "error": f"Estado civil '{estado_civil_raw}' no válido",
                        }
                    )
                    results["total_errors"] += 1
                    continue

                # Validar fecha de nacimiento
                fecha_nacimiento = row_data.get("Fecha de nacimiento")
                if not fecha_nacimiento:
                    results["errors"].append(
                        {
                            "row": row_num,
                            "error": "La fecha de nacimiento es obligatoria",
                        }
                    )
                    results["total_errors"] += 1
                    continue

                # Convertir fecha si es necesario
                if isinstance(fecha_nacimiento, str):
                    try:
                        fecha_nacimiento = datetime.strptime(
                            fecha_nacimiento, "%Y-%m-%d"
                        ).date()
                    except:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": "Formato de fecha de nacimiento inválido. Use YYYY-MM-DD",
                            }
                        )
                        results["total_errors"] += 1
                        continue
                else:
                    # Si es un objeto datetime de Excel
                    fecha_nacimiento = fecha_nacimiento.date()

                # Crear objeto usuario
                user_data = {
                    "nombres": str(row_data["Nombres"]).strip(),
                    "apellidos": str(row_data["Apellidos"]).strip(),
                    "n_documento": str(row_data["N° Documento"]).strip(),
                    "genero": genero,
                    "fecha_nacimiento": fecha_nacimiento,
                    "estado_civil": estado_civil,
                    "ocupacion_quedesempeño": str(row_data.get("Ocupación", "")).strip()
                    or "",
                    "tipo_usuario": tipo_usuario,
                    "visitas_domiciliarias": False,  # IMPORTANTE: Solo usuarios para fundación
                    "estado": "ACTIVO",
                    "escribe": False,
                    "lee": False,
                    "ha_estado_en_otro_centro": False,
                    "proteccion_exequial": False,
                    "is_deleted": False,
                    "fecha_registro": datetime.utcnow(),
                    "direccion": None,  # No requerido para usuarios de fundación
                    "telefono": None,  # No requerido para usuarios de fundación
                    "email": None,  # No requerido para usuarios de fundación
                    "nucleo_familiar": "Nuclear",
                    "grado_escolaridad": None,
                    "lugar_nacimiento": None,
                    "lugar_procedencia": None,
                    "origen_otrocentro": None,
                    "regimen_seguridad_social": None,
                    "tipo_afiliacion": None,
                    "profesion": str(row_data.get("Ocupación", "")).strip() or "",
                    "url_imagen": None,
                }

                # Crear usuario usando el CRUD existente
                user = User(**user_data)
                saved_user = crud.save_user(user, None)  # Sin foto

                results["success"].append(
                    {
                        "row": row_num,
                        "user_id": saved_user.id_usuario,
                        "nombre": f"{saved_user.nombres} {saved_user.apellidos}",
                    }
                )
                results["total_success"] += 1

            except Exception as e:
                # Hacer rollback de la sesión si hay error
                try:
                    crud.__carelink_session.rollback()
                except:
                    pass

                results["errors"].append(
                    {"row": row_num, "error": f"Error al procesar fila: {str(e)}"}
                )
                results["total_errors"] += 1

        return Response[dict](
            data=results,
            status_code=HTTPStatus.OK,
            message=f"Importación completada. {results['total_success']} usuarios creados, {results['total_errors']} errores.",
            error=None,
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}",
        )


@router.get("/family-members/template/excel")
async def export_family_member_template(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
):
    """Generar plantilla Excel para importación masiva de familiares"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        import tempfile
        import os

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Familiares"

        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Definir columnas
        columns = [
            "N° Documento del Paciente",
            "N° Documento del Familiar",
            "Nombres del Familiar",
            "Apellidos del Familiar",
            "Teléfono del Familiar",
            "Dirección del Familiar",
            "Email del Familiar",
            "Parentesco",
            "Es Acudiente (Sí/No)",
            "Vive (Sí/No)",
        ]

        # Agregar encabezados
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Ajustar ancho de columnas
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Agregar datos de ejemplo
        example_data = [
            [
                "12345678",
                "87654321",
                "María",
                "González",
                "3001234567",
                "Calle 123 #45-67",
                "maria@email.com",
                "Madre",
                "Sí",
                "Sí",
            ],
            [
                "23456789",
                "98765432",
                "Juan",
                "Pérez",
                "3009876543",
                "Carrera 78 #12-34",
                "juan@email.com",
                "Padre",
                "No",
                "Sí",
            ],
        ]

        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment

        # Guardar en archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name

        return FileResponse(
            tmp_file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="plantilla_familiares.xlsx",
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar plantilla: {str(e)}",
        )


@router.get("/usuarios/registrados")
async def list_all_authorized_users(
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(require_roles(Role.ADMIN.value)),
) -> Response[List[AuthorizedUserDTO]]:
    users_list = crud._get_all_authorized_users()
    users_list_response = [AuthorizedUserDTO.from_orm(user) for user in users_list]
    return Response[List[AuthorizedUserDTO]](
        message="Usuarios consultados",
        error=None,
        data=users_list_response,
        status_code=HTTPStatus.OK,
    )


@router.get("/usuarios/registrados/{user_id}")
async def list_authorized_user_by_id(
    user_id: int,
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(require_roles(Role.ADMIN.value)),
) -> Response[AuthorizedUserDTO]:
    user = crud._get_authorized_user_info(user_id)
    user_response = AuthorizedUserDTO.from_orm(user)
    return Response[AuthorizedUserDTO](
        message="Usuario consultado",
        error=None,
        status_code=HTTPStatus.OK,
        data=user_response,
    )


@router.post("/family-members/import/excel")
async def import_family_members_from_excel(
    file: UploadFile = File(...),
    crud: CareLinkCrud = Depends(get_crud),
    _: AuthorizedUsers = Depends(get_current_user),
):
    """Importar familiares desde archivo Excel"""
    try:
        from openpyxl import load_workbook
        import tempfile
        import os

        # Validar tipo de archivo
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Solo se permiten archivos Excel (.xlsx)",
            )

        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name

        try:
            # Cargar workbook
            wb = load_workbook(tmp_file_path, data_only=True)
            ws = wb.active

            # Validar columnas requeridas
            expected_columns = [
                "N° Documento del Paciente",
                "N° Documento del Familiar",
                "Nombres del Familiar",
                "Apellidos del Familiar",
                "Teléfono del Familiar",
                "Dirección del Familiar",
                "Email del Familiar",
                "Parentesco",
                "Es Acudiente (Sí/No)",
                "Vive (Sí/No)",
            ]

            headers = [cell.value for cell in ws[1]]
            if not all(col in headers for col in expected_columns):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="El archivo no contiene todas las columnas requeridas",
                )

            # Procesar filas
            results = {
                "success": [],
                "errors": [],
                "total_processed": 0,
                "total_success": 0,
                "total_errors": 0,
            }

            for row_num in range(2, ws.max_row + 1):
                results["total_processed"] += 1

                try:
                    # Obtener datos de la fila
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row_num, column=col).value
                        row_data[header] = cell_value if cell_value is not None else ""

                    # Validar datos requeridos
                    paciente_documento = str(
                        row_data.get("N° Documento del Paciente", "")
                    ).strip()
                    if not paciente_documento:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": "El número de documento del paciente es obligatorio",
                            }
                        )
                        results["total_errors"] += 1
                        continue

                    familiar_documento = str(
                        row_data.get("N° Documento del Familiar", "")
                    ).strip()
                    if not familiar_documento:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": "El número de documento del familiar es obligatorio",
                            }
                        )
                        results["total_errors"] += 1
                        continue

                    nombres = str(row_data.get("Nombres del Familiar", "")).strip()
                    if not nombres:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": "Los nombres del familiar son obligatorios",
                            }
                        )
                        results["total_errors"] += 1
                        continue

                    apellidos = str(row_data.get("Apellidos del Familiar", "")).strip()
                    if not apellidos:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": "Los apellidos del familiar son obligatorios",
                            }
                        )
                        results["total_errors"] += 1
                        continue

                    parentesco = str(row_data.get("Parentesco", "")).strip()
                    if not parentesco:
                        results["errors"].append(
                            {"row": row_num, "error": "El parentesco es obligatorio"}
                        )
                        results["total_errors"] += 1
                        continue

                    # Buscar paciente por documento
                    paciente = crud.get_user_by_document(paciente_documento)
                    if not paciente:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": f"No se encontró un paciente con el documento {paciente_documento}",
                            }
                        )
                        results["total_errors"] += 1
                        continue

                    # Validar si el familiar ya existe
                    existing_family_member = crud.get_family_member_by_document(
                        familiar_documento
                    )
                    if existing_family_member:
                        results["errors"].append(
                            {
                                "row": row_num,
                                "error": f"Ya existe un familiar con el documento {familiar_documento}",
                            }
                        )
                        results["total_errors"] += 1
                        continue

                    # Procesar campos opcionales
                    telefono = (
                        str(row_data.get("Teléfono del Familiar", "")).strip() or None
                    )
                    direccion = (
                        str(row_data.get("Dirección del Familiar", "")).strip() or None
                    )
                    email = str(row_data.get("Email del Familiar", "")).strip() or None

                    # Truncar campos según restricciones de la base de datos
                    if telefono and len(telefono) > 50:
                        telefono = telefono[:50]
                    if direccion and len(direccion) > 255:
                        direccion = direccion[:255]
                    if email and len(email) > 50:
                        email = email[:50]
                    if nombres and len(nombres) > 50:
                        nombres = nombres[:50]
                    if apellidos and len(apellidos) > 50:
                        apellidos = apellidos[:50]

                    # Validar email para tabla Usuarios (máximo 30 caracteres)
                    if email and len(email) > 30:
                        # Si el email es demasiado largo para Usuarios, no lo usamos
                        email = None

                    # Procesar campos booleanos
                    es_acudiente_raw = (
                        str(row_data.get("Es Acudiente (Sí/No)", "")).strip().lower()
                    )
                    es_acudiente = es_acudiente_raw in [
                        "sí",
                        "si",
                        "s",
                        "yes",
                        "y",
                        "true",
                        "1",
                    ]

                    vive_raw = str(row_data.get("Vive (Sí/No)", "")).strip().lower()
                    vive = vive_raw in ["sí", "si", "s", "yes", "y", "true", "1"]

                    # Verificar si ya existe un acudiente para este usuario
                    if es_acudiente:
                        existing_acudiente = crud.check_existing_acudiente(
                            paciente.id_usuario
                        )
                        if existing_acudiente:
                            results["errors"].append(
                                {
                                    "row": row_num,
                                    "error": f"El paciente ya tiene un acudiente registrado",
                                }
                            )
                            results["total_errors"] += 1
                            continue

                    # Crear datos del familiar
                    family_member_data = {
                        "n_documento": familiar_documento,
                        "nombres": nombres,
                        "apellidos": apellidos,
                        "telefono": telefono,
                        "direccion": direccion,
                        "email": email,
                        "acudiente": es_acudiente,
                        "vive": vive,
                        "is_deleted": False,
                    }

                    # Crear el familiar usando el método público del CRUD
                    family_member = crud.create_family_member(family_member_data)

                    # Crear la relación
                    relationship = crud.create_family_member_relationship(
                        paciente.id_usuario, family_member.id_acudiente, parentesco
                    )

                    # Actualizar campos del usuario si es acudiente
                    if es_acudiente:
                        crud.update_user_contact_info(paciente, family_member)

                    # Confirmar cambios
                    crud.commit_changes()

                    results["success"].append(
                        {
                            "row": row_num,
                            "family_member_id": family_member.id_acudiente,
                            "paciente_nombre": f"{paciente.nombres} {paciente.apellidos}",
                            "familiar_nombre": f"{nombres} {apellidos}",
                            "parentesco": parentesco,
                            "es_acudiente": es_acudiente,
                        }
                    )
                    results["total_success"] += 1

                except Exception as e:
                    # Hacer rollback de la sesión si hay error
                    try:
                        crud.__carelink_session.rollback()
                    except:
                        pass

                    results["errors"].append(
                        {"row": row_num, "error": f"Error al procesar fila: {str(e)}"}
                    )
                    results["total_errors"] += 1

            return Response[dict](
                data=results,
                status_code=HTTPStatus.OK,
                message=f"Importación completada. {results['total_success']} familiares creados, {results['total_errors']} errores.",
                error=None,
            )

        finally:
            # Limpiar archivo temporal
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}",
        )


@router.put("/authorized-users/{user_id}")
def update_authorized_user(
    user_id: int,
    update_data: AuthorizedUserUpdate,
    db: Session = Depends(get_carelink_db),
) -> AuthorizedUsers:
    user = db.query(AuthorizedUsers).filter(AuthorizedUsers.id == user_id).first()
    if not user:
        raise ValueError(f"User with ID {user_id} not found")

    professional = (
        db.query(Profesionales).filter(Profesionales.id_user == user_id).first()
    )

    user_fields = {"email", "first_name", "last_name", "role", "password"}

    professional_data = {
        "nombres": update_data.first_name or user.first_name,
        "apellidos": update_data.last_name or user.last_name,
        "n_documento": update_data.document_number,
        "t_profesional": update_data.professional_id_number,
        "fecha_nacimiento": update_data.birthdate,
        "fecha_ingreso": update_data.entry_date,
        "profesion": update_data.profession,
        "especialidad": update_data.specialty,
        "cargo": update_data.charge,
        "telefono": update_data.phone_number,
        "e_mail": update_data.email or user.email,
        "direccion": update_data.home_address,
        "estado": "Activo",
    }

    data = update_data.dict(exclude_unset=True)

    for field, value in data.items():
        if field in user_fields:
            if field == "password" and value:
                setattr(user, field, hash_password(value))
            else:
                setattr(user, field, value)

    # ✅ Handle role change
    new_role = data.get("role", user.role)

    if new_role != RoleEnum.profesional:
        if professional:
            professional.id_user = None
            db.add(professional)
    else:
        if any(v is not None for v in professional_data.values()):
            if professional:
                for field, value in professional_data.items():
                    if value is not None:
                        setattr(professional, field, value)
            else:
                new_professional = Profesionales(id_user=user_id, **professional_data)
                db.add(new_professional)

    db.commit()
    db.refresh(user)
    return user
